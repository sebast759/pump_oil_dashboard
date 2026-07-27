#!/usr/bin/env python3
"""
EU Weekly Oil Bulletin Dashboard Generator
==========================================
Reads the EC Weekly Oil Bulletin Excel file and produces a self-contained
HTML dashboard with:
  - 36-month historical price lines (EU, FR, DE, NL, ES, IT, PT)
  - YTD % change bars
  - Tax vs pre-tax breakdown
  - Consumption mix by country

Usage:
    python generate_oil_dashboard.py                     # refresh data and create site/index.html
    python generate_oil_dashboard.py --local             # site/index.html using cached data
    python generate_oil_dashboard.py --push              # generate and publish to GitHub Pages
    python generate_oil_dashboard.py path/to/file.xlsx   # use a specific Excel file
    python generate_oil_dashboard.py --local --output my.html
                                                        # custom local output file

Requirements:
    pip install openpyxl
"""

import os
import sys
import json
import math
import base64
import argparse
import tempfile
import shutil
import re
import html as html_lib
from pathlib import Path
from datetime import datetime, date, timezone, timedelta

try:
    import openpyxl
except ImportError:
    print("Missing dependency. Run:  pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------

# Controlled via --local CLI flag (see main()). Do not edit here.

DEFAULT_XLSX = "Weekly_Oil_Bulletin_Prices_History_maticni_4web.xlsx"

# These URLs are stable — the EC overwrites the same file in-place every Thursday.
# Verified: same UUID appears in official newsletter emails from Apr 2024 → Aug 2025.
HISTORY_URL = (
    "https://energy.ec.europa.eu/document/download/"
    "906e60ca-8b6a-44e7-8589-652854d2fd3f_en"
    "?filename=Weekly_Oil_Bulletin_Prices_History_maticni_4web.xlsx"
)
DUTIES_URL = (
    "https://energy.ec.europa.eu/document/download/"
    "ccdc6e96-6792-40cb-b0b4-b6609f1e30d0_en"
    "?filename=Oil_Bulletin_Duties_and_taxes.xlsx"
)


def download_file(url: str, dest: Path, label: str = "file") -> bool:
    """Download url to dest. Returns True on success."""
    try:
        import urllib.request
        print(f"  Downloading {label} ...")
        headers = {"User-Agent": "Mozilla/5.0 (compatible; oil-bulletin-dashboard/1.0)"}
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=60) as resp, open(dest, "wb") as fh:
            downloaded = 0
            while True:
                block = resp.read(65536)
                if not block:
                    break
                fh.write(block)
                downloaded += len(block)
        print(f"  Saved {downloaded/1024:.0f} KB -> {dest}")
        return True
    except Exception as e:
        print(f"  Download failed: {e}")
        return False


def resolve_xlsx(input_arg: str, force_download: bool, cache_dir: Path, local: bool = False) -> Path:
    """
    Resolution order:
      1. Explicit path that exists on disk  -> use it directly.
      2. --download flag                    -> always re-fetch from EC.
      3. Cached file < 7 days old           -> use cache.
      4. No file found                      -> auto-download.
    """
    # Explicit file path provided
    if input_arg != DEFAULT_XLSX:
        p = Path(input_arg)
        if p.exists():
            return p
        print(f"ERROR: File not found: {p}")
        sys.exit(1)

    cache_path = cache_dir / DEFAULT_XLSX
    cache_dir.mkdir(parents=True, exist_ok=True)

    if local:
        if cache_path.exists():
            print(f"  [LOCAL] Skipping download — using cached file -> {cache_path}")
            return cache_path
        print("ERROR: --local set but no cached xlsx found. Run once without --local.")
        sys.exit(1)

    if force_download:
        print("Force-downloading latest file from EC ...")
        if download_file(HISTORY_URL, cache_path, "Oil Bulletin history"):
            return cache_path
        sys.exit(1)

    # Check cache freshness (7 days)
    if cache_path.exists():
        age_days = (datetime.now() - datetime.fromtimestamp(cache_path.stat().st_mtime)).days
        if age_days < 5:
            print(f"  Using cached file (age: {age_days}d) -> {cache_path}")
            return cache_path
        else:
            print(f"  Cached file is {age_days} days old -- refreshing ...")

    # Auto-download
    if download_file(HISTORY_URL, cache_path, "Oil Bulletin history"):
        return cache_path

    # Fallback to stale cache if download failed
    if cache_path.exists():
        print("  WARNING: Download failed -- using stale cache.")
        return cache_path

    print("ERROR: No local file and download failed. Check your internet connection.")
    sys.exit(1)


BRENT_CACHE_CSV = Path(__file__).parent / "brent_cache.csv"


def _save_brent_cache(
    date_strs: list,
    aligned: list,
    brent_latest: dict | None,
    daily_dates: list[str] | None = None,
    daily_prices: list[float] | None = None,
):
    """Save aligned weekly and daily Brent prices to the offline cache."""
    import csv
    with open(BRENT_CACHE_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["date", "price"])
        for ds, price in zip(date_strs, aligned):
            w.writerow([ds, price if price is not None else ""])
        for ds, price in zip(daily_dates or [], daily_prices or []):
            w.writerow([f"daily:{ds}", price])
        if brent_latest:
            w.writerow([f"latest:{brent_latest['date']}", brent_latest["price"]])


def _load_brent_cache(date_strs: list) -> tuple:
    """Load Brent prices from CSV cache. Returns (aligned, brent_ytd, brent_latest)."""
    import csv
    if not BRENT_CACHE_CSV.exists():
        return [None] * len(date_strs), None, None
    brent_map = {}
    brent_latest = None
    daily_map = {}
    with open(BRENT_CACHE_CSV, newline="", encoding="utf-8") as f:
        for row in csv.reader(f):
            if len(row) < 2 or row[0] == "date":
                continue
            key, val = row[0], row[1]
            if not val:
                continue
            price = float(val)
            if key.startswith("latest:"):
                brent_latest = {"price": price, "date": key[7:]}
            elif key.startswith("daily:"):
                daily_map[key[6:]] = price
            else:
                brent_map[key] = price
    aligned = [brent_map.get(ds) for ds in date_strs]
    latest_yr = int(date_strs[-1][:4])
    jan1_str  = f"{latest_yr}-01-01"
    jan_idx   = max(0, next((i for i, ds in enumerate(date_strs) if ds >= jan1_str), 0) - 1)
    base = next((v for v in aligned[jan_idx:] if v), None)
    last = next((v for v in reversed(aligned) if v), None)
    brent_ytd = round((last / base - 1) * 100, 2) if base and last else None
    found = sum(1 for v in aligned if v is not None)
    print(f"  Brent (cache): {found}/{len(date_strs)} weeks loaded")
    daily_dates = sorted(daily_map)
    return aligned, brent_ytd, brent_latest, daily_dates, [daily_map[d] for d in daily_dates]


def fetch_brent(date_strs: list, local: bool = False) -> tuple:
    """
    Load official daily Brent spot data, extended with adjusted Yahoo BZ=F.

    EC weekly dates are aligned to the latest available trading observation on
    or before that date. Provider downloads are incrementally cached by
    ``brent_spot``; this legacy weekly cache remains the offline fallback.
    Returns (aligned_prices, brent_ytd_pct, brent_latest).
    """
    if local:
        print("  [LOCAL] Skipping FRED/Yahoo — loading Brent from cache ...")
        cached = _load_brent_cache(date_strs)
        if cached[3]:
            return cached
        try:
            import pandas as pd
            from brent_spot import get_continuous_brent_spot
            spot = get_continuous_brent_spot(
                start=date_strs[0],
                end=date.today(),
                cache_dir=Path(__file__).parent / ".cache" / "brent",
                cache_ttl=timedelta(days=36500),
                yahoo_cache_ttl=timedelta(days=36500),
            )
            return (*cached[:3], [d.strftime("%Y-%m-%d") for d in spot.index],
                    [round(float(v), 2) for v in spot])
        except Exception:
            return cached

    try:
        import pandas as pd
        from brent_spot import get_continuous_brent_spot

        start_dt = pd.Timestamp(date_strs[0]) - pd.Timedelta(days=14)
        spot = get_continuous_brent_spot(
            start=start_dt,
            end=date.today(),
            adjustment="constant",
            cache_dir=Path(__file__).parent / ".cache" / "brent",
            cache_ttl=timedelta(hours=6),
            yahoo_cache_ttl=timedelta(0),
            refresh_lookback_days=45,
        )

        aligned = []
        for ds in date_strs:
            value = spot.asof(pd.Timestamp(ds))
            aligned.append(round(float(value), 2) if pd.notna(value) else None)
        found   = sum(1 for v in aligned if v is not None)
        print(f"  Brent spot (FRED + adjusted Yahoo): {found}/{len(date_strs)} weeks matched")

        # YTD %
        latest_yr = int(date_strs[-1][:4])
        jan1_str  = f"{latest_yr}-01-01"
        jan_idx   = max(0, next((i for i, ds in enumerate(date_strs) if ds >= jan1_str), 0) - 1)
        base = next((v for v in aligned[jan_idx:] if v), None)
        last = next((v for v in reversed(aligned) if v), None)
        brent_ytd = round((last / base - 1) * 100, 2) if base and last else None

        latest_date = spot.index[-1]
        latest_price = float(spot.iloc[-1])
        brent_latest = {
            "price": round(latest_price, 2),
            "date": latest_date.strftime("%Y-%m-%d"),
        }
        print(
            f"  Brent latest continuous spot: "
            f"${latest_price:.2f} ({brent_latest['date']})"
        )

        daily_dates = [d.strftime("%Y-%m-%d") for d in spot.index]
        daily_prices = [round(float(v), 2) for v in spot]
        _save_brent_cache(
            date_strs, aligned, brent_latest, daily_dates, daily_prices
        )
        return aligned, brent_ytd, brent_latest, daily_dates, daily_prices

    except Exception as e:
        print(f"  WARNING: Brent download failed: {e} — using cache")
        return _load_brent_cache(date_strs)


COUNTRIES    = ["EU", "FR", "DE", "NL", "ES", "IT", "PT"]
COUNTRIES    = ["FR", "DE", "NL", "ES", "IT", "PT"]
WEEKS_BACK   = 522          # ≈ 10 years of weekly data

# Colours
COLORS = {
    "EU": "#f59e0b", "FR": "#3b82f6", "DE": "#10b981",
    "NL": "#f97316", "ES": "#ec4899", "IT": "#8b5cf6", "PT": "#06b6d4",
}
FUEL_COLORS = {
    "Gasoline":    "#f59e0b",
    "Diesel":      "#3b82f6",
    "Heating Oil": "#10b981",
    "Fuel Oil":    "#a1a1aa",
    "LPG":         "#f97316",
}

# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------
def safe_float(v):
    try:
        return float(v) if v is not None else None
    except (TypeError, ValueError):
        return None


def col_indices(headers, country, sheet_type="with_tax"):
    """Return dict {product: col_index} for a given country in a price sheet."""
    prefix_map = {
        "with_tax": f"{country}_price_with_tax_",
        "wo_tax":   f"{country}_price_wo_tax_",
        "cons":     f"{country}_consumption_",
    }
    prefix = prefix_map[sheet_type]
    result = {}
    for i, h in enumerate(headers):
        if h and str(h).startswith(prefix):
            # Normalise heating oil column name variants
            product = str(h)[len(prefix):]
            product = product.replace(f"he{country}ing_oil", "heating_oil")
            result[product] = i
    return result


def fmt_date(d):
    """datetime → 'Mon YY' label for chart axis."""
    months = ["","Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]
    return f"{months[d.month]} {str(d.year)[2:]}"


def fmt_date_full(d):
    """datetime → 'd Mon YY' label for chart axis (e.g. '5 Jan 26')."""
    months = ["","Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]
    return f"{d.day} {months[d.month]} {str(d.year)[2:]}"


def compute_sensitivity(brent: list, fuel: list, window: int = 4, lag: int = 1):
    """OLS slope through origin, split by direction of Brent move.

    Unit: millieuros/L per $/bbl = eurocents/L per $10/bbl (chart display unit).
    Returns (up_slope, down_slope).
    """
    up_xx = up_xy = dn_xx = dn_xy = 0.0
    for t in range(window + lag, len(brent)):
        bx  = brent[t - lag]
        bx0 = brent[t - lag - window]
        fy  = fuel[t]
        fy0 = fuel[t - window]
        if bx is None or bx0 is None or fy is None or fy0 is None:
            continue
        dx = bx - bx0   # $/bbl
        dy = fy  - fy0  # millieuros/L
        if   dx > 0: up_xx += dx * dx; up_xy += dx * dy
        elif dx < 0: dn_xx += dx * dx; dn_xy += dx * dy
    return (
        round(up_xy / up_xx, 2) if up_xx else 0.0,
        round(dn_xy / dn_xx, 2) if dn_xx else 0.0,
    )


# ---------------------------------------------------------------------------
# DATA EXTRACTION
# ---------------------------------------------------------------------------
def extract_data(xlsx_path: Path, local: bool = False) -> dict:
    print(f"Reading: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # ---- PRICES WITH TAX ------------------------------------------------
    ws_tax = wb["Prices with taxes"]
    tax_rows = list(ws_tax.iter_rows(values_only=True))
    tax_headers = tax_rows[0]

    # Data rows start at index 3 (index 0=col names, 1=labels, 2=units, 3+=data)
    data_rows = tax_rows[3:]
    # Filter to rows that have a valid date
    data_rows = [(r[0], r) for r in data_rows if isinstance(r[0], datetime)]
    # Sort oldest→newest
    data_rows.sort(key=lambda x: x[0])
    # Take last WEEKS_BACK
    data_rows = data_rows[-WEEKS_BACK:]

    dates            = [r[0] for _, r in data_rows]
    date_labels      = [fmt_date(d) for d in dates]
    date_labels_full = [fmt_date_full(d) for d in dates]
    date_strs        = [d.strftime("%Y-%m-%d") for d in dates]
    # Pump prices lag Brent by one week. The history chart therefore extends
    # one week beyond the pump series so the latest Brent observation remains
    # visible after being shifted to the following pump-price week.
    chart_dates       = dates + [dates[-1] + timedelta(days=7)]
    chart_labels      = [fmt_date(d) for d in chart_dates]
    chart_labels_full = [fmt_date_full(d) for d in chart_dates]

    # Column maps for each country
    col_tax = {c: col_indices(tax_headers, c, "with_tax") for c in COUNTRIES}

    # ---- PRICES WITHOUT TAX ----------------------------------------------
    ws_notax = wb["Prices wo taxes"]
    notax_rows_raw = list(ws_notax.iter_rows(values_only=True))
    notax_headers  = notax_rows_raw[0]
    notax_by_date  = {}
    for r in notax_rows_raw[3:]:
        if isinstance(r[0], datetime):
            notax_by_date[r[0].strftime("%Y-%m-%d")] = r

    col_notax = {c: col_indices(notax_headers, c, "wo_tax") for c in COUNTRIES}

    # ---- BUILD PRICE SERIES ----------------------------------------------
    countries_data = {}
    for c in COUNTRIES:
        euro95, diesel, euro95_nt, diesel_nt = [], [], [], []
        for d, r in data_rows:
            ds = d.strftime("%Y-%m-%d")
            # With tax
            v95  = safe_float(r[col_tax[c].get("euro95")])   if col_tax[c].get("euro95")  is not None else None
            vdie = safe_float(r[col_tax[c].get("diesel")])   if col_tax[c].get("diesel")  is not None else None
            euro95.append(round(v95,  2) if v95  is not None else None)
            diesel.append(round(vdie, 2) if vdie is not None else None)
            # Without tax
            nr = notax_by_date.get(ds)
            if nr:
                vnt95  = safe_float(nr[col_notax[c].get("euro95")])  if col_notax[c].get("euro95")  is not None else None
                vntdie = safe_float(nr[col_notax[c].get("diesel")])  if col_notax[c].get("diesel")  is not None else None
                euro95_nt.append(round(vnt95,  2) if vnt95  is not None else None)
                diesel_nt.append(round(vntdie, 2) if vntdie is not None else None)
            else:
                euro95_nt.append(None)
                diesel_nt.append(None)

        # Tax rate estimate from latest non-null pair
        tax_rate = None
        for i in range(len(euro95)-1, -1, -1):
            if euro95[i] and euro95_nt[i] and euro95[i] > 0:
                tax_rate = round(1 - euro95_nt[i] / euro95[i], 3)
                break

        countries_data[c] = {
            "euro95":       euro95,
            "diesel":       diesel,
            "euro95_notax": euro95_nt,
            "diesel_notax": diesel_nt,
            "tax_rate":     tax_rate or 0.55,
        }

    # ---- YTD % CHANGE ---------------------------------------------------
    ytd = {}
    jan1 = date(dates[-1].year, 1, 1)
    # Use last week of prior year as base (one step before the first date >= Jan 1)
    jan_idx = max(0, next((i for i, d in enumerate(dates) if d.date() >= jan1), 0) - 1)

    for c in COUNTRIES:
        s95  = countries_data[c]["euro95"]
        sdie = countries_data[c]["diesel"]
        def ytd_pct(series):
            base = next((v for v in series[jan_idx:] if v), None)
            last = next((v for v in reversed(series) if v), None)
            if base and last and base > 0:
                return round((last / base - 1) * 100, 2)
            return None
        def ytd_abs(series):
            base = next((v for v in series[jan_idx:] if v), None)
            last = next((v for v in reversed(series) if v), None)
            if base and last:
                return round((last - base) / 1000, 4)  # €/L
            return None
        ytd[c] = {
            "euro95_ytd": ytd_pct(s95), "diesel_ytd": ytd_pct(sdie),
            "euro95_abs": ytd_abs(s95), "diesel_abs": ytd_abs(sdie),
        }

    # ---- CONSUMPTION ----------------------------------------------------
    ws_cons  = wb["Consumption"]
    cons_rows = list(ws_cons.iter_rows(values_only=True))
    cons_headers = cons_rows[0]

    cons_ctrs = [c for c in COUNTRIES if c != "EU"]
    col_cons  = {c: col_indices(cons_headers, c, "cons") for c in cons_ctrs}

    # Latest year row (row index 2 after skipping headers)
    latest_cons_row = cons_rows[2]  # most recent year

    consumption = {}
    for c in cons_ctrs:
        vals = {}
        for prod, idx in col_cons[c].items():
            v = safe_float(latest_cons_row[idx])
            vals[prod] = round(v, 1) if v else 0.0
        # Rename for display
        consumption[c] = {
            "Gasoline":    vals.get("euro95",      0),
            "Diesel":      vals.get("diesel",       0),
            "Heating Oil": vals.get("heating_oil",  0),
            "Fuel Oil":    vals.get("fuel_oil_1",   0) + vals.get("fuel_oil_2", 0),
            "LPG":         vals.get("LPG",          0),
        }

    latest_year = latest_cons_row[0] if latest_cons_row[0] else "latest"

    (
        brent_prices,
        brent_ytd,
        brent_latest,
        brent_daily_dates,
        brent_daily_prices,
    ) = fetch_brent(date_strs, local=local)

    # Brent absolute YTD change ($/bbl)
    brent_jan_idx = max(0, next((i for i, ds in enumerate(date_strs) if ds >= f"{dates[-1].year}-01-01"), 0) - 1)
    brent_base    = next((v for v in brent_prices[brent_jan_idx:] if v), None)
    brent_last    = next((v for v in reversed(brent_prices) if v), None)
    brent_abs     = round(brent_last - brent_base, 2) if brent_base and brent_last else None

    # ---- SENSITIVITY (OLS, computed once here rather than in JS) -----------
    _SENS_WINDOWS = [4, 26]
    _SENS_LAGS    = [0, 1, 2]
    sensitivity   = {}
    for fuel_key in ("euro95", "diesel"):
        # Per-country slopes (default window=4, lag=1) for the bar charts
        per_country = {}
        for c in COUNTRIES:
            up, dn = compute_sensitivity(brent_prices, countries_data[c][fuel_key])
            per_country[c] = {"up": up, "down": dn}
        # Research grid: average slope across all countries for each (window, lag)
        research = {}
        for win in _SENS_WINDOWS:
            for lag in _SENS_LAGS:
                ups, dns = [], []
                for c in COUNTRIES:
                    u, d = compute_sensitivity(
                        brent_prices, countries_data[c][fuel_key], win, lag
                    )
                    ups.append(u); dns.append(d)
                research[f"{win}_{lag}"] = {
                    "up":   round(sum(ups) / len(ups), 2),
                    "down": round(sum(dns) / len(dns), 2),
                }
        sensitivity[fuel_key] = {"per_country": per_country, "research": research}

    print(f"  Dates: {date_strs[0]} to {date_strs[-1]} ({len(date_strs)} weeks)")
    print(f"  YTD ({dates[-1].year}): {ytd}")
    print(f"  Consumption year: {latest_year}")

    return {
        "dates":       date_strs,
        "labels":      date_labels,
        "labels_full": date_labels_full,
        "chart_labels":      chart_labels,
        "chart_labels_full": chart_labels_full,
        "countries":   countries_data,
        "ytd":         ytd,
        "consumption": consumption,
        "latest_year": str(latest_year),
        "latest_date": date_strs[-1],
        "brent":        brent_prices,
        "ytd_weeks":    len(date_strs) - jan_idx,
        "brent_ytd":    brent_ytd,
        "brent_abs":    brent_abs,
        "brent_latest": brent_latest,
        "brent_daily_dates": brent_daily_dates,
        "brent_daily": brent_daily_prices,
        "sensitivity":  sensitivity,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }


# ---------------------------------------------------------------------------
# HTML GENERATION
# ---------------------------------------------------------------------------
CHART_JS_CDN = "https://cdn.jsdelivr.net/npm/chart.js@4.4.2/dist/chart.umd.min.js"
SITE_ASSET_NAMES = (
    "og-image.png",
    "favicon.ico",
    "favicon-192.png",
    "favicon-512.png",
    "apple-touch-icon.png",
)
CONTENT_ASSET_NAMES = (
    "seb-dog-drive.jpg",
    "fuel-decision-illustration.png",
)


def emit_site_support_files(output_dir: Path) -> None:
    """Emit static SEO, discovery, and identity files beside the dashboard."""
    source_dir = Path(__file__).parent / "assets" / "site"
    for name in SITE_ASSET_NAMES:
        source = source_dir / name
        if not source.is_file():
            raise FileNotFoundError(f"Required site asset is missing: {source}")
        destination = output_dir / name
        if source.resolve() != destination.resolve():
            shutil.copy2(source, destination)

    for name in CONTENT_ASSET_NAMES:
        source = Path(__file__).parent / "assets" / name
        if not source.is_file():
            raise FileNotFoundError(f"Required content asset is missing: {source}")
        destination = output_dir / name
        if source.resolve() != destination.resolve():
            shutil.copy2(source, destination)

    (output_dir / "robots.txt").write_text(
        "User-agent: *\nAllow: /\n\nSitemap: https://fuelforecast.eu/sitemap.xml\n",
        encoding="utf-8",
    )
    (output_dir / "sitemap.xml").write_text(
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
        "  <url>\n"
        "    <loc>https://fuelforecast.eu/</loc>\n"
        f"    <lastmod>{date.today().isoformat()}</lastmod>\n"
        "    <changefreq>daily</changefreq>\n"
        "  </url>\n"
        "</urlset>\n",
        encoding="utf-8",
    )
    (output_dir / "site.webmanifest").write_text(
        json.dumps(
            {
                "name": "Fuel Forecast",
                "short_name": "Fuel Forecast",
                "start_url": "/",
                "display": "standalone",
                "background_color": "#131a2a",
                "theme_color": "#131a2a",
                "icons": [
                    {
                        "src": "/favicon-192.png",
                        "sizes": "192x192",
                        "type": "image/png",
                    },
                    {
                        "src": "/favicon-512.png",
                        "sizes": "512x512",
                        "type": "image/png",
                    },
                ],
            },
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )


def build_html(data: dict) -> str:
    colors_js      = json.dumps(COLORS)
    fuel_colors_js = json.dumps(FUEL_COLORS)
    countries_js   = json.dumps(COUNTRIES)
    data_js        = json.dumps(data)
    illustration_path = Path(__file__).parent / "assets" / "refuel-nozzle.png"
    refuel_illustration = base64.b64encode(illustration_path.read_bytes()).decode()
    goatcounter_code = (
        os.environ.get("GOATCOUNTER_SITE_CODE") or "sebast9"
    ).strip()
    if goatcounter_code and not re.fullmatch(r"[A-Za-z0-9-]+", goatcounter_code):
        raise ValueError(
            "GOATCOUNTER_SITE_CODE may contain only letters, numbers, and hyphens"
        )
    analytics_html = (
        f'<script data-goatcounter="https://{goatcounter_code}.goatcounter.com/count" '
        'async src="https://gc.zgo.at/count.js"></script>'
        if goatcounter_code else ""
    )

    previous_brent = next(
        (value for value in reversed(data["brent"][:-1]) if value is not None),
        None,
    )
    latest_brent = (
        data.get("brent_latest", {}).get("price")
        if data.get("brent_latest") else None
    )
    if latest_brent is None:
        latest_brent = next(
            (value for value in reversed(data["brent"]) if value is not None),
            None,
        )
    brent_move = (
        latest_brent - previous_brent
        if latest_brent is not None and previous_brent is not None else None
    )
    if brent_move is None:
        current_signal = (
            "The current signal is unavailable because the latest Brent move "
            "could not be calculated."
        )
    else:
        diesel_coefficient = 0.09 if brent_move >= 0 else 0.06
        expected_cents = diesel_coefficient * brent_move * 10
        rounded_cents = abs(round(expected_cents))
        if abs(expected_cents) < 2:
            current_signal = (
                "The current Diesel signal is NO RUSH because pump prices are "
                "expected to stay broadly flat next week."
            )
        elif expected_cents > 0:
            current_signal = (
                f"The current Diesel signal is GO NOW because pump prices could "
                f"rise by about {rounded_cents} cents/L next week."
            )
        else:
            current_signal = (
                f"The current Diesel signal is WAIT because pump prices could "
                f"fall by about {rounded_cents} cents/L next week."
            )
    current_signal += (
        " Fuel Forecast recomputes the signal each week from the latest "
        "available Brent move."
    )

    faq_entries = [
        (
            "Will fuel prices go up next week?",
            current_signal,
        ),
        (
            "How do rising oil prices affect pump prices?",
            "Crude oil moves first, then petrol and diesel prices adjust as "
            "stations restock. In the EU weekly data, the pass through appears "
            "with roughly a one week lag.",
        ),
        (
            "How much does a $10 Brent move change pump prices?",
            "The site’s Diesel estimates are about 9 cents/L when Brent rises "
            "and about 6 cents/L when Brent falls. This difference is known as "
            "asymmetric pass through.",
        ),
        (
            "How much can I save by timing a fill up?",
            "In normal weeks, timing a 50L fill up may save a few euros. The "
            "difference can be larger after a major crude move such as the "
            "2026 spike.",
        ),
        (
            "Where does the data come from?",
            "Pump prices come from the European Commission Weekly Oil Bulletin. "
            "Brent data comes from FRED and Yahoo Finance.",
        ),
        (
            "Is this a price prediction?",
            "No. It is an indicative signal from a simple pass through model, "
            "and local prices can differ.",
        ),
    ]
    faq_items = []
    for index, (question, answer) in enumerate(faq_entries):
        question_html = html_lib.escape(question)
        if index == 1:
            question_html = (
                f'<a href="#oil-price-pass-through" '
                'onclick="showTab(4);setTimeout(()=&gt;'
                "document.getElementById('oil-price-pass-through')"
                ".scrollIntoView({behavior:'smooth'}),0)\">"
                f"{question_html}</a>"
            )
        answer_html = html_lib.escape(answer)
        if index == 4:
            answer_html = (
                'Pump prices come from the '
                '<a href="https://energy.ec.europa.eu/data-and-analysis/'
                'weekly-oil-bulletin_en" target="_blank" '
                'rel="noopener noreferrer">European Commission Weekly Oil '
                'Bulletin</a>. Brent data comes from '
                '<a href="https://fred.stlouisfed.org/series/DCOILBRENTEU" '
                'target="_blank" rel="noopener noreferrer">FRED</a> and '
                '<a href="https://finance.yahoo.com/quote/BZ%3DF/" '
                'target="_blank" rel="noopener noreferrer">Yahoo Finance</a>.'
            )
        faq_items.append(
            f'<div class="faq-item"><h3>{question_html}</h3>'
            f'<p>{answer_html}</p></div>'
        )
    faq_html = "\n".join(faq_items)
    faq_schema_js = json.dumps(
        {
            "@type": "FAQPage",
            "@id": "https://fuelforecast.eu/#faq",
            "mainEntity": [
                {
                    "@type": "Question",
                    "name": question,
                    "acceptedAnswer": {
                        "@type": "Answer",
                        "text": answer,
                    },
                }
                for question, answer in faq_entries
            ],
        },
        ensure_ascii=False,
    )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Fuel Forecast · Petrol and Diesel Prices in Europe</title>
<meta name="description" content="Weekly diesel and Euro-95 pump prices across Europe, with a next-week forecast from Brent crude moves. Should you fill up now or wait?">
<meta name="robots" content="index, follow">
<meta name="author" content="Fuel Forecast">
<meta name="theme-color" content="#131a2a">
<link rel="canonical" href="https://fuelforecast.eu/">
<meta property="og:type" content="website">
<meta property="og:site_name" content="Fuel Forecast">
<meta property="og:locale" content="en_GB">
<meta property="og:url" content="https://fuelforecast.eu/">
<meta property="og:title" content="Fuel Forecast · Petrol and Diesel Prices in Europe">
<meta property="og:description" content="Weekly diesel and Euro-95 pump prices across Europe, with a next-week forecast from Brent crude moves. Should you fill up now or wait?">
<meta property="og:image" content="https://fuelforecast.eu/og-image.png">
<meta property="og:image:width" content="1200">
<meta property="og:image:height" content="630">
<meta property="og:image:alt" content="Fuel Forecast next-week outlook and European pump-price chart">
<meta name="twitter:card" content="summary_large_image">
<meta name="twitter:title" content="Fuel Forecast · Petrol and Diesel Prices in Europe">
<meta name="twitter:description" content="Weekly diesel and Euro-95 pump prices across Europe, with a next-week forecast from Brent crude moves. Should you fill up now or wait?">
<meta name="twitter:image" content="https://fuelforecast.eu/og-image.png">
<meta name="twitter:image:alt" content="Fuel Forecast next-week outlook and European pump-price chart">
<script type="application/ld+json">
{{
  "@context": "https://schema.org",
  "@graph": [
    {{
      "@type": "WebSite",
      "@id": "https://fuelforecast.eu/#website",
      "name": "Fuel Forecast",
      "url": "https://fuelforecast.eu/",
      "description": "Weekly European pump prices with a practical next-week fuel-price outlook.",
      "inLanguage": "en"
    }},
    {{
      "@type": "Dataset",
      "@id": "https://fuelforecast.eu/#dataset",
      "name": "Fuel Forecast European Pump Price and Brent Dataset",
      "description": "A derived presentation of weekly European diesel and Euro-95 pump prices alongside daily Brent spot prices.",
      "url": "https://fuelforecast.eu/",
      "isAccessibleForFree": true,
      "license": "https://commission.europa.eu/legal-notice_en",
      "creator": {{"@type": "Organization", "name": "Fuel Forecast"}},
      "variableMeasured": ["Diesel pump price", "Euro-95 pump price", "Brent spot price"],
      "isBasedOn": [
        {{
          "@type": "Dataset",
          "name": "European Commission Weekly Oil Bulletin",
          "url": "https://energy.ec.europa.eu/data-and-analysis/weekly-oil-bulletin_en"
        }},
        {{
          "@type": "Dataset",
          "name": "FRED Crude Oil Prices: Brent - Europe (DCOILBRENTEU)",
          "url": "https://fred.stlouisfed.org/series/DCOILBRENTEU"
        }}
      ]
    }},
    {faq_schema_js}
  ]
}}
</script>
<link rel="icon" href="favicon.ico" sizes="any">
<link rel="icon" type="image/png" sizes="192x192" href="favicon-192.png">
<link rel="icon" type="image/png" sizes="512x512" href="favicon-512.png">
<link rel="apple-touch-icon" sizes="180x180" href="apple-touch-icon.png">
<link rel="manifest" href="site.webmanifest">
<script src="{CHART_JS_CDN}"></script>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=DM+Sans:wght@400;500;600;700;800&family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
:root {{
  --bg-page: #131a2a;
  --bg-card: #1a2336;
  --bg-inset: #212b42;
  --border: rgba(148,163,184,.12);
  --text-primary: #e2e8f0;
  --text-secondary: #94a3b8;
}}
body {{
  background: var(--bg-page);
  color: var(--text-primary);
  font-family: 'DM Sans', sans-serif;
  min-height: 100vh;
}}
.mono {{ font-family: 'DM Mono', monospace; }}
/* HEADER */
.header {{
  background: linear-gradient(135deg,#182236 0%,#1d2a42 100%);
  border-bottom: 1px solid var(--border);
  padding: 20px 32px 0;
}}
.header-top {{
  display: flex; align-items: flex-start;
  justify-content: space-between; margin-bottom: 16px; flex-wrap: wrap; gap: 12px;
}}
.header-title {{ display: flex; align-items: center; gap: 14px; }}
.logo {{
  width: 82px; height: 58px; flex: 0 0 82px;
  display: flex; align-items: center; justify-content: center;
}}
.logo img {{
  display:block; width:100%; height:100%; object-fit:contain;
}}
h1 {{ font-size: 24px; font-weight: 800; letter-spacing: -0.5px; color: #f8fafc; }}
.subtitle {{ font-size: 13px; color: #b6c5d8; margin-top: 3px; }}
.price-badges {{ display: flex; gap: 8px; flex-wrap: wrap; justify-content: flex-end; }}
.badge {{
  background: var(--bg-inset); border-radius: 8px;
  padding: 6px 12px; text-align: center; min-width: 68px;
}}
.badge-label {{ font-size: 10px; font-weight: 700; margin-bottom: 2px; }}
.badge-price {{ font-size: 13px; font-weight: 800; color: #e2e8f0; }}
.badge-unit {{ font-size: 9px; color: #94a3b8; margin-top: 1px; }}
/* TABS */
.tabs {{ display: flex; }}
.tab-btn {{
  background: none; border: none; cursor: pointer;
  padding: 10px 20px; font-size: 13px; font-weight: 600;
  color: #94a3b8; border-bottom: 2px solid transparent;
  transition: all 0.2s; font-family: 'DM Sans', sans-serif;
}}
.tab-btn.active {{ color: #f59e0b; border-bottom-color: #f59e0b; }}
/* PANELS */
.content {{ padding: 24px 32px; max-width: 1100px; margin: 0 auto; }}
.panel {{ display: none; }}
.panel.active {{ display: block; }}
/* CARDS / BOXES */
.card {{
  background: var(--bg-card); border-radius: 12px;
  border: 1px solid var(--border); overflow: hidden;
}}
.card-header {{
  padding: 14px 20px; border-bottom: 1px solid var(--border);
  display: flex; justify-content: space-between; align-items: center;
}}
.card-title {{ font-size: 13px; font-weight: 700; color: #e2e8f0; }}
.card-sub {{ font-size: 11px; color: #94a3b8; }}
/* GRID */
.grid-2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
.grid-7 {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(110px, 1fr)); gap: 10px; margin-bottom: 20px; }}
/* YTD cards */
.ytd-card {{
  background: var(--bg-card); border-radius: 10px;
  border: 1px solid var(--border); padding: 12px;
}}
.ytd-ctr {{ font-size: 11px; font-weight: 700; margin-bottom: 8px; }}
.ytd-val {{ font-size: 13px; font-weight: 800; line-height: 1.3; }}
.ytd-sub {{ font-size: 9px; color: #94a3b8; margin-bottom: 2px; }}
.ytd-val2 {{ font-size: 12px; font-weight: 700; margin-top: 4px; line-height: 1.3; }}
.up {{ color: #f87171; }}
.dn {{ color: #34d399; }}
/* TABLE */
table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
th {{
  padding: 8px 16px; text-align: right; color: #94a3b8;
  font-weight: 600; border-bottom: 1px solid var(--border);
  background: var(--bg-inset);
}}
th:first-child {{ text-align: left; }}
td {{ padding: 8px 16px; text-align: right; }}
td:first-child {{ text-align: left; }}
tr:nth-child(even) {{ background: rgba(33,43,66,.48); }}
.dot {{
  display: inline-block; width: 8px; height: 8px;
  border-radius: 50%; margin-right: 8px;
}}
.pill {{
  background: #f59e0b20; color: #f59e0b;
  padding: 2px 8px; border-radius: 12px;
  font-size: 11px;
}}
/* FUEL toggle */
.toggle-row {{ display: flex; gap: 8px; }}
.toggle-btn {{
  padding: 6px 14px; border-radius: 20px; font-size: 12px;
  font-weight: 600; border: 1px solid var(--border);
  background: var(--bg-inset); color: #94a3b8;
  cursor: pointer; font-family: 'DM Sans', sans-serif;
  transition: all 0.2s;
}}
.toggle-btn.active {{
  border-color: #f59e0b; background: rgba(245,158,11,0.1); color: #f59e0b;
}}
.history-section {{
  margin-top:30px; padding-top:28px;
  border-top:1px solid rgba(148,163,184,.22);
}}
.chart-controls {{
  width:max-content; max-width:100%; margin:0 auto;
  padding:13px 18px 12px; border-radius:10px;
  background:var(--bg-card); border:1px solid var(--border);
}}
.chart-controls-title {{
  color:#64748b; font-size:9px; font-weight:800;
  letter-spacing:.12em; text-align:center; margin-bottom:9px;
}}
.chart-control-row {{
  display:grid; grid-template-columns:72px auto;
  gap:10px; align-items:center; margin-top:8px;
}}
.chart-control-label {{
  color:#94a3b8; font-size:10px; font-weight:700; text-align:right;
}}
/* TAX bars */
.tax-bar-row {{
  display: flex; align-items: center; gap: 12px; margin-bottom: 10px;
}}
.tax-bar-label {{ width: 32px; font-size: 12px; font-weight: 700; }}
.tax-bar-track {{
  flex: 1; background: var(--bg-inset); border-radius: 4px; height: 26px; overflow: hidden;
}}
.tax-bar-fill {{
  height: 100%; border-radius: 4px;
  display: flex; align-items: center; justify-content: flex-end; padding-right: 10px;
}}
.tax-bar-pct {{ font-size: 11px; font-weight: 700; color: #fff; }}
.tax-bar-aside {{ width: 100px; font-size: 11px; color: #94a3b8; text-align: right; }}
/* Chart containers */
.chart-wrap {{ padding: 16px 12px 8px; }}
.history-legend {{
  display: flex; flex-direction: column; align-items: center; gap: 7px;
  padding: 14px 16px 0; font-size: 11px;
}}
.history-legend-row {{
  display: flex; align-items: center; justify-content: center;
  gap: 10px; flex-wrap: wrap;
}}
.history-legend-group {{
  color: #64748b; font-size: 9px; font-weight: 800;
  letter-spacing: .08em; text-transform: uppercase; min-width: 68px;
}}
.history-legend-item {{
  display: inline-flex; align-items: center; gap: 5px;
  color: #94a3b8; background: none; border: 0; padding: 2px;
  font: inherit;
}}
button.history-legend-item {{ cursor: pointer; transition: opacity .15s, color .15s; }}
.history-legend-item.dimmed {{ opacity: .25; }}
.history-legend-item.focused {{ color: #e2e8f0; font-weight: 700; }}
.history-legend-item.hidden {{ opacity: .3; text-decoration: line-through; }}
.history-swatch {{ display: inline-block; width: 15px; height: 8px; border: 2px solid; }}
.history-line {{ display: inline-block; width: 18px; border-top: 2px dashed; }}
.history-diamond {{
  display: inline-block; width: 7px; height: 7px;
  border: 2px solid #94a3b8; transform: rotate(45deg);
}}
.refuel-callout {{
  position: relative; width: 100%; text-align: center; margin: 0 0 24px;
  padding: 25px 28px 23px; border-radius: 12px;
  background: linear-gradient(135deg, rgba(33,43,66,.96), rgba(26,35,54,.98));
  border: 1px solid var(--border);
  box-shadow: inset 0 1px 0 rgba(255,255,255,.025);
  font-family: 'Inter', 'DM Sans', sans-serif;
}}
.refuel-illustration {{
  position:absolute; left:18px; top:50%; transform:translateY(-50%);
  width:280px; height:180px; object-fit:contain; opacity:.92;
}}
.refuel-copy {{ padding:2px 150px 2px 150px; }}
.refuel-question {{
  color:#f8fafc; font-size:19px; font-weight:600;
  letter-spacing:-.015em; margin-bottom:9px;
}}
.refuel-answer {{
  font-weight:800; line-height:1.12; letter-spacing:-.025em;
  transition: font-size .2s ease, color .2s ease;
}}
.refuel-action, .refuel-detail {{ display: block; }}
.refuel-detail {{
  color:#dbe5f3; font-size:17px; font-weight:500;
  line-height:1.4; margin-top:7px; letter-spacing:-.01em;
}}
.refuel-context {{
  color:#dbe5f3; font-size:13px; font-weight:450;
  line-height:1.5; margin-top:11px;
}}
.refuel-context-line {{ display: block; }}
.saving-line {{
  width:max-content; max-width:100%; margin:0 auto;
  padding:5px 11px; border-radius:999px;
  color:#ecfdf5; background:rgba(16,185,129,.12);
  border:1px solid rgba(16,185,129,.25); font-weight:600;
}}
.market-line {{ margin-top:13px; color:#cbd5e1; font-weight:350; }}
.market-line strong {{ color:#f8fafc; font-weight:500; }}
@media (max-width: 700px) {{
  .refuel-illustration {{ width:160px; height:110px; left:4px; opacity:.25; }}
  .refuel-copy {{ padding: 2px 12px; position: relative; }}
  .chart-controls {{ width:100%; }}
  .chart-control-row {{ grid-template-columns:1fr; justify-items:center; }}
  .chart-control-label {{ text-align:center; }}
}}
canvas {{ max-width: 100%; }}
/* Info box */
.info-box {{
  margin-top: 12px; padding: 12px 16px;
  background: var(--bg-inset); border-radius: 8px;
  font-size: 11px; color: #94a3b8;
}}
.section-title {{
  font-size: 16px; font-weight: 700; color: #e2e8f0; margin-bottom: 4px;
}}
.section-sub {{ font-size: 12px; color: #94a3b8; margin-bottom: 20px; }}
/* About & sources */
.about-hero {{
  padding: 28px; margin-bottom: 20px;
  background: linear-gradient(135deg, rgba(33,43,66,.96), rgba(26,35,54,.98));
  border: 1px solid var(--border); border-radius: 12px;
}}
.about-hero-layout {{
  display:grid; grid-template-columns:340px minmax(0,1fr);
  gap:26px; align-items:center;
}}
.about-hero-copy {{ grid-column:2; grid-row:1; }}
.about-hero h2 {{ color:#f8fafc; font-size:24px; margin-bottom:10px; }}
.about-hero p {{ color:#cbd5e1; font-size:11pt; line-height:1.7; max-width:820px; }}
.about-hero-image {{
  display:block; width:100%; aspect-ratio:4/3; object-fit:cover;
  border-radius:10px; border:1px solid var(--border);
  grid-column:1; grid-row:1;
}}
.about-grid {{
  display:grid; grid-template-columns:repeat(3,1fr); gap:16px; margin-bottom:20px;
}}
.source-card {{
  display:block; padding:18px; min-height:150px; color:inherit;
  background:var(--bg-card); border:1px solid var(--border); border-radius:10px;
  text-decoration:none; transition:border-color .18s ease, transform .18s ease;
}}
.source-card:hover {{ border-color:#f59e0b; transform:translateY(-2px); }}
.source-kicker {{
  color:#60a5fa; font-size:10pt; font-weight:800;
  letter-spacing:.12em; text-transform:uppercase; margin-bottom:8px;
}}
.source-title {{ color:#f8fafc; font-size:13pt; font-weight:800; margin-bottom:7px; }}
.source-copy {{ color:#cbd5e1; font-size:11pt; line-height:1.6; }}
.source-link {{ color:#f59e0b; font-size:10.5pt; font-weight:700; margin-top:12px; }}
.about-copy {{ padding:20px; }}
.about-copy h3 {{ color:#e2e8f0; font-size:13pt; margin-bottom:8px; }}
.about-copy p, .about-copy li {{ color:#cbd5e1; font-size:11pt; line-height:1.7; }}
.about-copy ul {{ padding-left:18px; }}
.panel#tab5 > .section-title {{ font-size:14pt; }}
.about-site {{
  margin-top:20px; padding:22px;
  background:var(--bg-card); border:1px solid var(--border); border-radius:10px;
}}
.about-site h3 {{ color:#f8fafc; font-size:16px; margin-bottom:10px; }}
.about-site-layout {{
  display:grid; grid-template-columns:minmax(0,1fr) 260px;
  gap:24px; align-items:center;
}}
.about-site p {{ color:#cbd5e1; font-size:11pt; line-height:1.7; max-width:880px; }}
.about-site-photo {{
  display:block; width:100%; aspect-ratio:1/1; object-fit:cover;
  border-radius:10px; border:1px solid var(--border);
}}
.about-contact {{ margin-top:15px; color:#94a3b8; font-size:11pt; line-height:1.8; }}
.about-contact a {{ color:#f59e0b; text-decoration:none; font-weight:700; }}
.about-contact a:hover {{ text-decoration:underline; }}
.faq-section {{
  margin-top:20px; padding:22px;
  background:var(--bg-card); border:1px solid var(--border); border-radius:10px;
}}
.faq-section > h2 {{ color:#f8fafc; font-size:18pt; margin-bottom:4px; }}
.faq-item {{ padding:16px 0; border-bottom:1px solid var(--border); }}
.faq-item:last-child {{ padding-bottom:0; border-bottom:0; }}
.faq-item h3 {{ color:#e2e8f0; font-size:13pt; margin-bottom:7px; }}
.faq-item p {{ color:#cbd5e1; font-size:11pt; line-height:1.7; }}
.faq-item a {{ color:#f59e0b; text-decoration:none; }}
.faq-item a:hover {{ text-decoration:underline; }}
/* Responsive */
@media (max-width: 900px) {{
  .grid-2 {{ grid-template-columns: 1fr; }}
  .grid-7 {{ grid-template-columns: repeat(4,1fr); }}
  .about-grid {{ grid-template-columns:1fr; }}
  .price-badges {{ display: none; }}
  .content {{ padding: 16px; }}
  .header {{ padding: 16px 16px 0; }}
  .about-hero-layout {{ grid-template-columns:1fr; }}
  .about-hero-copy {{ grid-column:auto; grid-row:auto; }}
  .about-hero-image {{
    width:min(100%,520px); margin:0 auto;
    grid-column:auto; grid-row:auto;
  }}
  .about-site-layout {{ grid-template-columns:1fr; }}
  .about-site-photo {{ width:min(100%,360px); margin:0 auto; }}
}}
</style>
</head>
<body>

<div class="header">
  <div class="header-top">
    <div>
      <div class="header-title">
        <div class="logo">
          <img src="data:image/png;base64,{refuel_illustration}"
               alt="Fuel Forecast nozzle logo">
        </div>
        <div>
          <h1>Fuel Forecast</h1>
          <div class="subtitle">Petrol and diesel prices in Europe · Should you fill up now or wait?</div>
          <button type="button" onclick="showTab(5)"
                  style="margin-top:5px;padding:0;border:0;background:none;color:#f59e0b;font:600 11px 'Inter','DM Sans',sans-serif;cursor:pointer;">
            About &amp; data sources →
          </button>
        </div>
      </div>
    </div>
    <div style="text-align:right;">
      <div style="display:flex;align-items:center;gap:10px;justify-content:flex-end;">
        <div id="badge-date-label" style="font-size:11px;color:#94a3b8;white-space:nowrap;line-height:1.4;"></div>
        <div class="price-badges" id="price-badges"></div>
      </div>
      <div id="brent-info"  style="font-size:10px;color:#94a3b8;margin-top:6px;"></div>
      <div id="brent-info2" style="font-size:10px;color:#94a3b8;margin-top:2px;"></div>
    </div>
  </div>
  <div class="tabs">
    <button class="tab-btn active" onclick="showTab(0)">Prices &amp; Forecast</button>
    <button class="tab-btn" onclick="showTab(1)" id="ytd-tab-label">2026 Variation</button>
    <button class="tab-btn" onclick="showTab(2)">How Much Tax in 1 Litre?</button>
    <button class="tab-btn" onclick="showTab(3)">Consumption Mix: Diesel or SP95?</button>
    <button class="tab-btn" onclick="showTab(4)">Sensitivity to Brent Price</button>
    <button class="tab-btn" onclick="showTab(5)">About &amp; Sources</button>
  </div>
</div>

<div class="content">

  <!-- TAB 0: Historical -->
  <div class="panel active" id="tab0">
    <div style="margin-bottom:20px;">
      <div class="refuel-callout">
        <img class="refuel-illustration"
             src="data:image/png;base64,{refuel_illustration}"
             alt="" aria-hidden="true">
        <div class="refuel-copy">
          <div class="refuel-question">Should I refuel my car today?</div>
          <div class="refuel-answer" id="refuel-answer"></div>
          <div class="refuel-context" id="refuel-context"></div>
        </div>
      </div>
      <div class="history-section">
        <div class="section-title" style="text-align:center;">How have pump prices changed?</div>
        <div class="section-sub" style="text-align:center;margin-bottom:14px;">Weekly consumer pump prices inclusive of taxes and duties</div>
        <div class="chart-controls">
          <div class="chart-controls-title">CHART OPTIONS</div>
          <div class="chart-control-row">
            <div class="chart-control-label">Fuel shown</div>
            <div class="toggle-row" aria-label="Fuel shown on the chart">
              <button class="toggle-btn" id="btn95" onclick="switchFuel('euro95')">Euro-95</button>
              <button class="toggle-btn active" id="btnD" onclick="switchFuel('diesel')">Diesel</button>
            </div>
          </div>
          <div class="chart-control-row">
            <div class="chart-control-label">Time range</div>
            <div class="toggle-row" aria-label="Time range shown on the chart">
              <button class="toggle-btn active" id="btnYTD" onclick="setRange(DATA.ytd_weeks)">YTD</button>
              <button class="toggle-btn" id="btn1Y" onclick="setRange(52)">1Y</button>
              <button class="toggle-btn" id="btn3Y" onclick="setRange(156)">3Y</button>
              <button class="toggle-btn" id="btn5Y" onclick="setRange(260)">5Y</button>
              <button class="toggle-btn" id="btnAll" onclick="setRange(0)">ALL</button>
            </div>
          </div>
        </div>
        <div style="display:flex;flex-direction:column;align-items:center;">
          <div class="section-sub" id="history-status" style="margin-top:10px;"></div>
        </div>
      </div>
    </div>
    <div class="card">
      <div class="history-legend" id="history-legend"></div>
      <div class="chart-wrap" style="height:500px;">
        <canvas id="histChart"></canvas>
      </div>
    </div>
    <div class="card" style="margin-top:20px;">
      <div class="card-header">
        <span class="card-title">Latest Prices — <span id="table-date"></span></span>
        <span class="card-sub">€ / litre</span>
      </div>
      <div class="chart-wrap" style="height:300px;">
        <canvas id="priceChart"></canvas>
      </div>
    </div>
  </div>

  <!-- TAB 1: YTD -->
  <div class="panel" id="tab1">
    <div class="section-title">How have prices changed in <span id="ytd-year"></span>?</div>
    <div class="section-sub">Change from Jan 1 through latest data point</div>
    <div class="grid-7" id="ytd-cards"></div>
    <div id="ytd-insight"></div>
    <div class="card">
      <div class="card-header"><span class="card-title">YTD Price Change</span><span class="card-sub">€/L for pump prices · $/bbl for Brent</span></div>
      <div class="chart-wrap" style="height:340px;">
        <canvas id="ytdAbsChart"></canvas>
      </div>
    </div>
    <div class="card" style="margin-top:16px;">
      <div class="card-header"><span class="card-title">YTD Price Change %</span><span class="card-sub">Relative to Jan 1</span></div>
      <div class="chart-wrap" style="height:360px;">
        <canvas id="ytdChart"></canvas>
      </div>
    </div>
  </div>

  <!-- TAB 2: Tax Analysis -->
  <div class="panel" id="tab2">
    <div class="section-title">How Much Tax Is in 1 Litre?</div>
    <div class="section-sub">Pre-tax product cost vs total duty burden · EUR/litre · Latest data point</div>
    <div class="grid-2">
      <div class="card">
        <div class="card-header"><span class="card-title">Diesel — Pre-tax vs Taxes (€/L)</span></div>
        <div class="chart-wrap" style="height:280px;"><canvas id="taxDChart"></canvas></div>
        <div style="padding:0 16px 16px;" id="taxDTable"></div>
      </div>
      <div class="card">
        <div class="card-header"><span class="card-title">Euro-95 — Pre-tax vs Taxes (€/L)</span></div>
        <div class="chart-wrap" style="height:280px;"><canvas id="tax95Chart"></canvas></div>
        <div style="padding:0 16px 16px;" id="tax95Table"></div>
      </div>
    </div>
    <div class="card" style="margin-top:20px;">
      <div class="card-header"><span class="card-title">Effective Tax Rate on Euro-95 (ranked)</span></div>
      <div style="padding:16px 20px;" id="tax-bars"></div>
    </div>
  </div>

  <!-- TAB 3: Consumption -->
  <div class="panel" id="tab3">
    <div style="margin-bottom:20px;">
      <div class="section-title" style="text-align:center;">Which fuels do European countries use? · <span id="cons-year"></span></div>
      <div class="section-sub" style="text-align:center;margin-bottom:14px;">Source: EC Weekly Oil Bulletin · kt (1,000 tonnes)</div>
      <div style="display:flex;justify-content:center;">
        <div class="toggle-row">
          <button class="toggle-btn active" id="btnAbsolute" onclick="switchCons('absolute')">By Volume (kt)</button>
          <button class="toggle-btn" id="btnMix" onclick="switchCons('mix')">Consumption Mix %</button>
        </div>
      </div>
    </div>
    <!-- Sub-panel: absolute volumes -->
    <div id="cons-absolute">
      <div class="card">
        <div class="chart-wrap" style="height:400px;"><canvas id="consAbsChart"></canvas></div>
      </div>
    </div>
    <!-- Sub-panel: percentage mix -->
    <div id="cons-mix" style="display:none;">
      <div class="card">
        <div class="chart-wrap" style="height:400px;"><canvas id="consMixChart"></canvas></div>
      </div>
    </div>
    <!-- Shared table -->
    <div class="card" style="margin-top:20px;">
      <div class="card-header"><span class="card-title">Absolute Volumes (kt)</span></div>
      <table>
        <thead><tr id="cons-head"></tr></thead>
        <tbody id="cons-body"></tbody>
      </table>
    </div>
  </div>

  <!-- TAB 4: Sensitivity -->
  <div class="panel" id="tab4">
    <div style="margin-bottom:20px;">
      <div class="section-title" id="oil-price-pass-through" style="text-align:center;">How do rising oil prices affect pump prices?</div>
      <div class="section-sub" style="text-align:center;margin:6px 0 2px;">
        Crude oil moves first, and the pass through to petrol and diesel prices appears as stations restock.
      </div>
      <div style="text-align:center;font-size:20px;font-weight:300;color:#f59e0b;margin:6px 0 2px;">
        Suppliers are <span style="text-decoration:underline;">quick</span> to pass on crude oil price increases (~9 cents for every $10 rise),<br>but <span style="text-decoration:underline;">slow</span> to adjust downward (~6 cents for every $10 decline)
      </div>
    </div>
    <div class="section-sub" style="text-align:center;margin-bottom:0;">
        Pump price rise / decline (€/L) per $10 Brent move
    </div>
  
    <div class="grid-2">
      <div>
        <div class="card">
          <div class="card-header">
            <span class="card-title">Diesel</span>
            <span class="card-sub" id="sens-diesel-slope"></span>
          </div>
          <div class="chart-wrap" style="height:340px;"><canvas id="sensDieselChart"></canvas></div>
        </div>
        <div class="card" style="margin-top:12px;" id="sens-diesel-table"></div>
      </div>
      <div>
        <div class="card">
          <div class="card-header">
            <span class="card-title">Euro-95</span>
            <span class="card-sub" id="sens-95-slope"></span>
          </div>
          <div class="chart-wrap" style="height:340px;"><canvas id="sens95Chart"></canvas></div>
        </div>
        <div class="card" style="margin-top:12px;" id="sens-95-table"></div>
      </div>
    </div>
    <div class="info-box" style="margin-top:16px;">
      <strong style="color:#f1f5f9;">How to read:</strong> Each bar = OLS slope of 4-week fuel change vs 4-week Brent change (lagged 1 week),
      split by direction of Brent move. <strong style="color:#f59e0b;">Solid bar</strong> = weeks when Brent rose ·
      <strong style="color:#94a3b8;">Light bar</strong> = weeks when Brent fell.
      A taller solid bar than light bar confirms the well-known asymmetry:
      <em>suppliers are quick to pass on crude price increases but slow to adjust downward</em>.
    </div>
    <div style="margin-top:24px;">
      <div class="section-title" style="text-align:center;margin-bottom:4px;">Why does the model use this time window and lag?</div>
      <div style="text-align:center;font-size:12px;color:#94a3b8;margin-bottom:14px;">
        Pump price rise / decline (€/L) per $10 Brent move &nbsp;·&nbsp; avg across all countries &nbsp;·&nbsp;
        <strong style="color:#f59e0b;">★ = current model choice (4W, lag 1W)</strong>
      </div>
      <div class="grid-2">
        <div class="card">
          <div class="card-header"><span class="card-title">Diesel</span></div>
          <div id="sens-research-diesel"></div>
        </div>
        <div class="card">
          <div class="card-header"><span class="card-title">Euro-95</span></div>
          <div id="sens-research-95"></div>
        </div>
      </div>
      <div class="info-box" style="margin-top:16px;font-size:12px;line-height:1.7;">
        <strong style="color:#f1f5f9;display:block;margin-bottom:6px;">Why 4-week window &amp; 1-week lag?</strong>
        <ul style="margin:0;padding-left:18px;color:#94a3b8;">
          <li>
            <strong style="color:#cbd5e1;">4-week window (short-run pass-through).</strong>
            Borenstein, Cameron &amp; Gilbert (1997, <em>QJE</em>) establish the 4-week horizon as
            the standard for measuring asymmetric retail-fuel price adjustment — long enough to
            capture a full pricing cycle, short enough to isolate behavioural response from
            seasonal and demand-side noise. Confirmed in the ECB's cross-country replication
            (Gelos &amp; Ustyugova, 2017) and in Aucremanne &amp; Dhyne (2005) for euro-area retail prices.
          </li>
          <li style="margin-top:6px;">
            <strong style="color:#cbd5e1;">1-week lag on Brent.</strong>
            EC prices are collected on Mondays and reflect purchasing decisions from the prior
            week. A 1-week lag aligns the crude signal with the actual procurement window, as
            recommended by Granger &amp; Lee (1989, <em>Oxford Bulletin of Econ &amp; Stats</em>) for
            error-correction models with weekly data, and applied in Bacon (1991,
            <em>Oxford Energy Studies</em>) to EU petrol markets.
          </li>
          <li style="margin-top:6px;">
            <strong style="color:#cbd5e1;">Why not 26W?</strong>
            Longer windows accumulate multiple Brent cycles; the growing gap above reflects
            compounded short-run asymmetry, not a stronger behavioural effect. With fewer
            independent observations and greater exposure to structural breaks (COVID, Ukraine),
            the 26W slope is less reliable for identifying pricing behaviour
            (cf. Peltzman 2000, <em>Journal of Political Economy</em>).
          </li>
        </ul>
      </div>
    </div>

  </div>

  <!-- TAB 5: About & Sources -->
  <div class="panel" id="tab5">
    <div class="about-hero">
      <div class="about-hero-layout">
        <div class="about-hero-copy">
          <h2>Fuel Forecast: should you refuel now or wait a week?</h2>
          <p>
            Crude moves first, pump prices follow with a lag of about a week.
            This tool tracks that lag across Europe and turns the latest Brent move
            into one practical signal: fill up now, or wait.
          </p>
          <p style="margin-top:12px;color:#93c5fd;font-weight:700;">
            Official EU pump prices, weekly. Brent spot, daily.
          </p>
        </div>
        <img class="about-hero-image" src="fuel-decision-illustration.png"
             alt="Driver considering whether to refuel beside a petrol pump"
             loading="lazy">
      </div>
    </div>

    <div class="about-site">
      <h3>About this site</h3>
      <div class="about-site-layout">
        <div>
          <p>
            My name is Seb and I live in Portugal. When the Iran war started on a
            Saturday, gasoline was at €1.50 and I drove straight to the station ⛽,
            expecting a queue. There was none. I filled up alone.
          </p>
          <p style="margin-top:10px;">
            Over the following days Brent jumped from $70 to almost $140 📈 and pump prices from €1.60 to
            €2.30. Stations take a few days to catch up with crude, and as a driver
            that lag is your opportunity 💡.
          </p>
          <p style="margin-top:10px;">
            This website is as simple as that: how much do you save on a full tank
            if you fill up this Friday instead of waiting until Tuesday next week ❓
          </p>
          <p style="margin-top:6px;font-weight:700;color:#f8fafc;">
            We give you the answer.
          </p>
          <div class="about-contact">
            Contact · <span id="site-contact">smaillard75 [at] gmail [dot] com</span>
          </div>
        </div>
        <img class="about-site-photo" src="seb-dog-drive.jpg"
             alt="Dog looking out of a car window during a drive in Portugal"
             loading="lazy">
      </div>
    </div>

    <div class="grid-2" style="margin-top:20px;">
      <div class="card about-copy">
        <h3>What we watch</h3>
        <ul>
          <li>What drivers pay for diesel and SP95 across Europe each week.</li>
          <li>How Brent crude moves from one day to the next.</li>
          <li>How prices, taxes and fuel habits differ between countries.</li>
          <li>Fresh Brent prices whenever the dashboard updates.</li>
        </ul>
      </div>
      <div class="card about-copy">
        <h3>How to use the signal</h3>
        <p>
          If Brent jumps, filling up now could help you get ahead of the next
          pump-price rise. If Brent falls, waiting a week could save you money.
          If very little has changed, there is no rush either way.
        </p>
        <p style="margin-top:10px;">
          Think of it as a helpful heads-up, not a promise. Your local price can
          still move differently.
        </p>
      </div>
    </div>

    <div class="section-title" style="margin:20px 0 12px;">Where does the data come from?</div>
    <div class="about-grid">
      <a class="source-card"
         href="https://energy.ec.europa.eu/data-and-analysis/weekly-oil-bulletin_en"
         target="_blank" rel="noopener noreferrer">
        <div class="source-kicker">Pump prices</div>
        <div class="source-title">EU Commission Weekly Oil Bulletin</div>
        <div class="source-copy">
          Weekly consumer prices including and excluding taxes, plus duties and
          petroleum-consumption data for European countries.
        </div>
        <div class="source-link">Open official source ↗</div>
      </a>
      <a class="source-card"
         href="https://fred.stlouisfed.org/series/DCOILBRENTEU"
         target="_blank" rel="noopener noreferrer">
        <div class="source-kicker">Brent Crude Oil · History</div>
        <div class="source-title">FRED · DCOILBRENTEU</div>
        <div class="source-copy">
          Daily Europe Brent spot price in US dollars per barrel. This is the
          preferred Brent series wherever an official observation is available.
        </div>
        <div class="source-link">View the FRED series ↗</div>
      </a>
      <a class="source-card"
         href="https://finance.yahoo.com/quote/BZ%3DF/"
         target="_blank" rel="noopener noreferrer">
        <div class="source-kicker">Brent Crude Oil · Real Time</div>
        <div class="source-title">Yahoo Finance · BZ=F</div>
        <div class="source-copy">
          Brent futures observations used to fill unpublished FRED trading dates
          and extend the series beyond FRED's latest release after level adjustment.
        </div>
        <div class="source-link">View the Yahoo quote ↗</div>
      </a>
    </div>

    <section class="faq-section" aria-labelledby="faq-heading">
      <h2 id="faq-heading">Fuel Price Questions</h2>
      {faq_html}
    </section>
  </div>

</div>

<div style="text-align:center;padding:16px 32px;font-size:12px;color:#94a3b8;border-top:1px solid #1e293b;">
  Fuel Forecast · Generated <span id="gen-datetime"></span> · Sources: EU Oil Bulletin, FRED &amp; Yahoo ·
  <a id="footer-contact" href="#" style="color:#f59e0b;text-decoration:none;">contact me</a>
</div>

<script>
// ---- EMBEDDED DATA -------------------------------------------------------
const DATA = {data_js};
const COLORS = {colors_js};
const FUEL_COLORS = {fuel_colors_js};
const CTRS = {countries_js};
const CHART_GRID = 'rgba(148,163,184,0.08)';
const TOOLTIP_BG = '#212b42';
const BRENT_COLOR = '#d4c98a';
const FUEL_TYPES = ["Gasoline","Diesel","Heating Oil","Fuel Oil","LPG"];
const FUEL_DISPLAY = {{"Gasoline":"SP95","Diesel":"Diesel","Heating Oil":"Heating Oil","Fuel Oil":"Fuel Oil","LPG":"LPG"}};

// ---- UTILS ---------------------------------------------------------------
const $ = id => document.getElementById(id);
let histChart, ytdChart, tax95Chart, taxDChart, consAbsChart, consMixChart;
let currentFuel  = 'diesel';
let currentRange = DATA.ytd_weeks;
let currentCons  = 'absolute';

function fmtVal(v, decimals=1) {{
  if (v == null) return '—';
  return v.toLocaleString('en', {{minimumFractionDigits: decimals, maximumFractionDigits: decimals}});
}}

// ---- INIT ----------------------------------------------------------------
document.addEventListener('DOMContentLoaded', () => {{
  const footerContact = $('footer-contact');
  footerContact.href = ['mai', 'lto:'].join('') +
    ['smaillard75', 'gmail.com'].join('@');
  const latest = DATA.dates[DATA.dates.length - 1];
  $('table-date').textContent  = latest;
  $('history-status').textContent =
    `Observed through ${{fmtDateLabel(latest)}} · Indicative forecast ${{DATA.chart_labels_full[DATA.dates.length]}}`;
  updateRefuelCallout();

  buildBadges();
  buildHistChart();
  setRange(DATA.ytd_weeks);
  buildPriceChart();
  buildYTD();
  buildTaxCharts();
  buildConsumption();
  buildSensitivity();
  $('gen-datetime').textContent = DATA.generated_at;
}});

let historyResizeTimer;
window.addEventListener('resize', () => {{
  clearTimeout(historyResizeTimer);
  historyResizeTimer = setTimeout(() => {{
    if (histChart) updateHistChart();
  }}, 120);
}});

function showTab(n) {{
  document.querySelectorAll('.panel').forEach((p,i) => p.classList.toggle('active', i===n));
  document.querySelectorAll('.tab-btn').forEach((b,i) => b.classList.toggle('active', i===n));
}}

// ---- BADGES (header) -----------------------------------------------------
function fmtDateLabel(iso) {{
  const d = new Date(iso + 'T12:00:00Z');
  return d.toLocaleDateString('en-GB', {{ day: 'numeric', month: 'long', year: 'numeric', timeZone: 'UTC' }});
}}

function brentInfoLine(price, date, diffPrice, diffRef, col) {{
  const sign  = diffPrice >= 0 ? '+' : '';
  const arrow = diffPrice >= 0 ? '▲' : '▼';
  const pct   = (diffPrice / diffRef * 100).toFixed(1);
  return `<span class="mono" style="color:#cbd5e1">$${{price.toFixed(2)}}/bbl</span>`
    + ` <span style="color:#94a3b8">(${{date}})</span>`
    + ` &nbsp;·&nbsp; <span style="color:${{col}}">${{arrow}} ${{sign}}$${{Math.abs(diffPrice).toFixed(2)}} (${{sign}}${{pct}}%)</span>`;
}}

function buildBadges() {{
  const last       = DATA.dates.length - 1;
  const latestDate = DATA.dates[last];
  const wrap       = $('price-badges');

  // Date label to the left of badges
  $('badge-date-label').innerHTML = fmtDateLabel(latestDate).replace(' ', '<br>');

  // Country badges (no per-badge date)
  CTRS.forEach(c => {{
    const v  = DATA.countries[c].euro95[last];
    const el = document.createElement('div');
    el.className = 'badge';
    el.style.border = `1px solid ${{COLORS[c]}}40`;
    el.innerHTML = `
      <div class="badge-label" style="color:${{COLORS[c]}}">${{c}}</div>
      <div class="badge-price mono">€${{v != null ? (v/1000).toFixed(3) : '—'}}</div>
      <div class="badge-unit">Euro-95/L</div>`;
    wrap.appendChild(el);
  }});

  // Brent badge: use the actual latest Yahoo observation.
  const brentVal = DATA.brent_latest?.price ?? DATA.brent[last];
  const bEl = document.createElement('div');
  bEl.className = 'badge';
  bEl.style.border = '1px solid #94a3b840';
  bEl.innerHTML = `
    <div class="badge-label" style="color:#94a3b8">Brent</div>
    <div class="badge-price mono" style="color:#94a3b8">${{brentVal != null ? '$'+brentVal.toFixed(2) : '—'}}</div>
    <div class="badge-unit">$/barrel</div>`;
  wrap.appendChild(bEl);

  // Latest daily Yahoo observation versus the preceding weekly observation.
  const bl = DATA.brent_latest;
  const cur = bl?.price ?? DATA.brent[last];
  const curDate = bl?.date ?? latestDate;
  let prev = null;
  for (let i = DATA.brent.length - 2; i >= 0; i--) {{
    if (DATA.brent[i] != null) {{ prev = DATA.brent[i]; break; }}
  }}
  if (cur != null && prev != null) {{
    const col = (cur - prev) >= 0 ? '#f87171' : '#34d399';
    $('brent-info').innerHTML =
      `Brent: ` + brentInfoLine(cur, curDate, cur - prev, prev, col)
      + ` <span style="color:#94a3b8">vs prev. week</span>`;
  }}

  $('brent-info2').innerHTML = '';
}}

// ---- HISTORICAL CHART ----------------------------------------------------
function latestBrentMove() {{
  const previous = DATA.brent.slice(0, -1).reverse().find(v => v != null);
  const latest = DATA.brent_latest?.price ?? DATA.brent[DATA.brent.length - 1];
  return latest != null && previous != null ? latest - previous : null;
}}

const DAY_MS = 24 * 60 * 60 * 1000;

function dateX(iso, leadDays=0) {{
  return Date.parse(iso + 'T12:00:00Z') + leadDays * DAY_MS;
}}

function displayedBrentSeries(startDate=DATA.dates[0]) {{
  if (DATA.brent_daily_dates?.length) {{
    return DATA.brent_daily_dates
      .map((date, index) => ({{
        x: dateX(date, 7),
        y: DATA.brent_daily[index],
        sourceDate: date
      }}))
      .filter(point => point.y != null && point.sourceDate >= startDate);
  }}
  return DATA.dates
    .map((date, index) => ({{
      x: dateX(date, 7),
      y: DATA.brent[index],
      sourceDate: date
    }}))
    .filter(point => point.y != null && point.sourceDate >= startDate);
}}

function forecastCoefficientForFuel(fuel, brentMove) {{
  if (fuel === 'diesel')
    return brentMove >= 0 ? 0.09 : 0.06;
  return brentMove >= 0 ? 0.07 : 0.05;
}}

function forecastCoefficient(brentMove) {{
  return forecastCoefficientForFuel(currentFuel, brentMove);
}}

function updateRefuelCallout() {{
  const move = latestBrentMove();
  const answer = $('refuel-answer');
  const context = $('refuel-context');
  if (move == null) {{
    answer.innerHTML = '<span class="refuel-action">NO SIGNAL</span>';
    answer.style.fontSize = '22px';
    answer.style.color = '#94a3b8';
    context.textContent = 'Latest Brent movement is unavailable';
    return;
  }}

  const coefficient = forecastCoefficient(move);
  const expectedCents = coefficient * move * 10;
  const cents = Math.abs(Math.round(expectedCents));
  const tankSaving = (Math.round(Math.abs(expectedCents / 100 * 50) * 10) / 10).toFixed(2);
  const magnitude = Math.abs(move);
  const strength = Math.max(0, Math.min(1, (magnitude - 3) / 7));
  answer.style.fontSize = `${{(22 + strength * 10).toFixed(1)}}px`;
  const weeklyBrent = [...DATA.brent].reverse().find(value => value != null);
  const currentBrent = DATA.brent_latest?.price ?? weeklyBrent;
  const brentDirection = move >= 0 ? 'up' : 'down';
  const expectedEuroLabel =
    `${{expectedCents >= 0 ? '+' : '-'}}€${{Math.abs(expectedCents / 100).toFixed(2)}}/L`;
  const brentLine =
    `<span class="refuel-context-line market-line">` +
    `Brent is ${{brentDirection}} <strong>$${{Math.abs(move).toFixed(1)}} this week</strong> ` +
    `at <strong>$${{currentBrent.toFixed(0)}}/bbl</strong>. Pump prices will follow next week ` +
    `(<strong>${{expectedEuroLabel}} expected</strong>).</span>`;

  if (Math.abs(expectedCents) < 2) {{
    answer.innerHTML =
      '<span class="refuel-action">NO RUSH</span>' +
      '<span class="refuel-detail">Pump prices should stay broadly flat next week</span>';
    answer.style.color = '#94a3b8';
    context.innerHTML = brentLine;
  }} else if (expectedCents > 0) {{
    answer.innerHTML =
      `<span class="refuel-action">GO NOW</span>` +
      `<span class="refuel-detail">Pump prices could rise ~+${{cents}} cents/L next week</span>`;
    answer.style.color = '#34d399';
    context.innerHTML =
      `<span class="refuel-context-line saving-line">That’s €${{tankSaving}} saved on a 50L tank</span>` +
      brentLine;
  }} else {{
    answer.innerHTML =
      `<span class="refuel-action">WAIT</span>` +
      `<span class="refuel-detail">Pump prices could fall ~${{cents}} cents/L next week</span>`;
    answer.style.color = '#f59e0b';
    context.innerHTML =
      `<span class="refuel-context-line saving-line">Waiting could save about €${{tankSaving}} on a 50L tank</span>` +
      brentLine;
  }}
}}

function pumpForecastForFuel(country, fuel) {{
  const series = DATA.countries[country][fuel];
  const latest = [...series].reverse().find(v => v != null);
  const brentMove = latestBrentMove();
  if (latest == null || brentMove == null) return null;
  const coefficient = forecastCoefficientForFuel(fuel, brentMove);
  return +(latest / 1000 + coefficient * brentMove / 10).toFixed(4);
}}

function pumpForecast(country) {{
  return pumpForecastForFuel(country, currentFuel);
}}

function pumpHistoryWithForecast(country) {{
  const points = DATA.countries[country][currentFuel]
    .map((value, index) => value == null ? null : ({{
      x: dateX(DATA.dates[index]),
      y: +(value / 1000).toFixed(4),
      sourceDate: DATA.dates[index]
    }}))
    .filter(Boolean);
  const forecast = pumpForecast(country);
  const forecastDisplayX = Math.max(
    dateX(DATA.dates[DATA.dates.length - 1], 7),
    DATA.brent_latest?.date
      ? dateX(DATA.brent_latest.date, 7)
      : dateX(DATA.dates[DATA.dates.length - 1], 7)
  );
  if (forecast != null) points.push({{
    x: forecastDisplayX,
    y: forecast,
    sourceDate: DATA.dates[DATA.dates.length - 1],
    forecast: true
  }});
  return points;
}}

function sharedPumpAxisBounds(start) {{
  // One shared vertical grid: Brent / 50 is its pump-price equivalent.
  // This makes $100/bbl and €2.00/L occupy exactly the same pixel row,
  // with every $10 Brent interval aligned to a €0.20/L pump interval.
  const values = [];
  for (const country of CTRS) {{
    for (const fuel of ['diesel', 'euro95']) {{
      DATA.countries[country][fuel].slice(start).forEach(value => {{
        if (value != null) values.push(value / 1000);
      }});
      const forecast = pumpForecastForFuel(country, fuel);
      if (forecast != null) values.push(forecast);
    }}
  }}
  const startDate = DATA.dates[start];
  DATA.brent_daily_dates.forEach((date, index) => {{
    const value = DATA.brent_daily[index];
    if (date >= startDate && value != null) values.push(value / 50);
  }});
  if (DATA.brent_latest?.price != null)
    values.push(DATA.brent_latest.price / 50);
  if (!values.length) return null;
  const low = Math.min(...values);
  const high = Math.max(...values);
  // A small buffer is enough because bounds are then rounded outward to a
  // full €0.20 / $10 grid interval. The epsilon prevents binary floating
  // point noise from adding an unnecessary extra interval.
  const padding = Math.max(0.02, (high - low) * 0.01);
  const epsilon = 1e-9;
  const min = Math.floor((low - padding + epsilon) / 0.20) * 0.20;
  const max = Math.ceil((high + padding - epsilon) / 0.20) * 0.20;
  return {{
    min,
    max,
    brentMin: min * 50,
    brentMax: max * 50
  }};
}}

function historyRightEdge(startDate) {{
  // Convert a fixed visual margin into the equivalent x-axis range. This
  // keeps the final labels about 150 px from the right axis for every range.
  const targetPixels = 150;
  const chartWidth =
    histChart?.chartArea?.width ||
    Math.max(500, ($('histChart').clientWidth || 900) - 120);
  const startX = dateX(startDate);
  const latestBrentX = DATA.brent_latest?.date
    ? dateX(DATA.brent_latest.date, 7)
    : dateX(DATA.dates[DATA.dates.length - 1], 7);
  const forecastX = dateX(DATA.dates[DATA.dates.length - 1], 7);
  const finalPointX = Math.max(latestBrentX, forecastX);
  const plottedWidth = Math.max(250, chartWidth - targetPixels);
  return finalPointX + (finalPointX - startX) * targetPixels / plottedWidth;
}}

function refreshHistoryLegendFocus() {{
  if (!histChart) return;
  const focus = histChart.$hoverCountry;
  document.querySelectorAll('.history-legend-country').forEach(button => {{
    const index = +button.dataset.index;
    const visible = histChart.isDatasetVisible(index);
    button.classList.toggle('focused', focus === index);
    button.classList.toggle('hidden', !visible);
    button.classList.toggle('dimmed', visible && focus != null && focus !== index);
    button.setAttribute('aria-pressed', visible ? 'true' : 'false');
  }});
}}

function setHistoryHover(index) {{
  if (!histChart || histChart.$hoverCountry === index) return;
  histChart.$hoverCountry = index;
  histChart.draw();
  refreshHistoryLegendFocus();
}}

function toggleHistoryCountry(index) {{
  const visible = histChart.isDatasetVisible(index);
  histChart.setDatasetVisibility(index, !visible);
  if (histChart.$hoverCountry === index) histChart.$hoverCountry = null;
  histChart.update();
  refreshHistoryLegendFocus();
}}

function buildHistoryLegend() {{
  const container = $('history-legend');
  container.innerHTML = `
    <div class="history-legend-row">
      <span class="history-legend-group">Countries</span>
      ${{CTRS.map((country, index) => `
        <button class="history-legend-item history-legend-country" data-index="${{index}}" type="button">
          <span class="history-swatch" style="border-color:${{COLORS[country]}}"></span>${{country}}
        </button>`).join('')}}
    </div>
    <div class="history-legend-row">
      <span class="history-legend-group">References</span>
      <span class="history-legend-item">
        <span class="history-line" style="border-color:#d4c98a"></span>Brent (1-week lead)
      </span>
      <span class="history-legend-item">
        <span class="history-line" style="border-color:#94a3b8"></span>
        <span class="history-diamond"></span>Indicative pump forecast
      </span>
    </div>`;

  container.querySelectorAll('.history-legend-country').forEach(button => {{
    const index = +button.dataset.index;
    button.addEventListener('mouseenter', () =>
      setHistoryHover(histChart.isDatasetVisible(index) ? index : null));
    button.addEventListener('mouseleave', () => setHistoryHover(null));
    button.addEventListener('click', () => toggleHistoryCountry(index));
  }});
}}

function buildHistChart() {{
  const ctx = $('histChart').getContext('2d');
  const datasets = CTRS.map(c => ({{
    label: c,
    hidden: ['DE', 'NL', 'IT'].includes(c),
    data: pumpHistoryWithForecast(c),
    borderColor: COLORS[c],
    backgroundColor: 'transparent',
    borderWidth: 1.8,
    pointStyle: ctx => ctx.dataIndex === ctx.dataset.data.length - 1 ? 'rectRot' : 'circle',
    pointRadius: ctx => ctx.dataIndex === ctx.dataset.data.length - 1 ? 5
      : (ctx.dataIndex === ctx.dataset.data.length - 2 ? 3 : 0),
    pointHoverRadius: ctx => ctx.dataIndex === ctx.dataset.data.length - 1 ? 7
      : (ctx.dataIndex === ctx.dataset.data.length - 2 ? 6 : 3),
    pointBackgroundColor: ctx => ctx.dataIndex === ctx.dataset.data.length - 1 ? '#1a2336' : COLORS[c],
    pointBorderColor: COLORS[c],
    pointBorderWidth: 2,
    segment: {{
      borderDash: ctx => ctx.p1DataIndex === ctx.chart.data.datasets[ctx.datasetIndex].data.length - 1
        ? [7, 4] : undefined,
      borderWidth: ctx => ctx.p1DataIndex === ctx.chart.data.datasets[ctx.datasetIndex].data.length - 1
        ? 2 : undefined
    }},
    tension: 0.15,
    spanGaps: true,
    yAxisID: 'y',
  }}));

  // Brent crude — left axis (rose)
  datasets.push({{
    label: 'Brent daily (1-week lead)',
    // Daily observations retain their real source date in tooltips but are
    // positioned seven days to the right to show their pump-price lead.
    data: displayedBrentSeries(),
    borderColor: BRENT_COLOR,
    backgroundColor: 'transparent',
    borderWidth: 1.5,
    borderDash: [5, 4],
    pointRadius: ctx => ctx.dataIndex === ctx.dataset.data.length - 1 ? 4 : 0,
    pointHoverRadius: ctx => ctx.dataIndex === ctx.dataset.data.length - 1 ? 6 : 3,
    pointBackgroundColor: BRENT_COLOR,
    pointBorderColor: '#1a2336',
    pointBorderWidth: 2,
    tension: 0,
    spanGaps: true,
    yAxisID: 'y2',
  }});

  datasets.push({{
    label: 'Pump forecast (1-week lag)',
    data: [],
    borderColor: '#94a3b8',
    backgroundColor: 'transparent',
    borderWidth: 1.5,
    borderDash: [6, 4],
    pointStyle: 'rectRot',
    pointBackgroundColor: '#1a2336',
    pointBorderColor: '#94a3b8',
  }});

  histChart = new Chart(ctx, {{
    type: 'line',
    data: {{ datasets }},
    options: {{
      responsive: true, maintainAspectRatio: false, devicePixelRatio: window.devicePixelRatio,
      interaction: {{ mode: 'nearest', axis: 'xy', intersect: false }},
      onHover: (event, elements, chart) => {{
        const hovered = elements.length && CTRS.includes(chart.data.datasets[elements[0].datasetIndex]?.label)
          ? elements[0].datasetIndex : null;
        if (chart.$hoverCountry !== hovered) {{
          chart.$hoverCountry = hovered;
          chart.draw();
          refreshHistoryLegendFocus();
        }}
      }},
      plugins: {{
        legend: {{
          display: false,
          labels: {{ color: '#94a3b8', font: {{ size: 11 }}, boxWidth: 16 }},
          onHover: (event, item, legend) => {{
            const chart = legend.chart;
            const hovered = CTRS.includes(chart.data.datasets[item.datasetIndex]?.label)
              ? item.datasetIndex : null;
            if (chart.$hoverCountry !== hovered) {{
              chart.$hoverCountry = hovered;
              chart.draw();
            }}
          }},
          onLeave: (event, item, legend) => {{
            if (legend.chart.$hoverCountry != null) {{
              legend.chart.$hoverCountry = null;
              legend.chart.draw();
            }}
          }},
          onClick: (event, item, legend) => {{
            const chart = legend.chart;
            if (!CTRS.includes(chart.data.datasets[item.datasetIndex]?.label)) {{
              Chart.defaults.plugins.legend.onClick(event, item, legend);
              return;
            }}
            chart.$lockedCountry = chart.$lockedCountry === item.datasetIndex
              ? null : item.datasetIndex;
            chart.draw();
          }}
        }},
        tooltip: {{
          backgroundColor: TOOLTIP_BG, borderColor: 'rgba(148,163,184,.2)', borderWidth: 1,
          titleColor: '#94a3b8', bodyColor: '#e2e8f0',
          xAlign: 'left', yAlign: 'center',
          titleAlign: 'left', bodyAlign: 'left',
          titleFont: {{ size: 10, weight: '600' }},
          bodyFont: {{ size: 10, weight: '400' }},
          padding: 8, bodySpacing: 3, boxWidth: 8, boxHeight: 8,
          callbacks: {{
            title: items => {{
              if (!items.length) return '';
              const item = items[0];
              const isLatestBrent =
                item.dataset.label.startsWith('Brent') &&
                item.dataIndex === item.dataset.data.length - 1;
              if (isLatestBrent && DATA.brent_latest?.date)
                return fmtDateLabel(DATA.brent_latest.date);
              const isPumpForecast =
                CTRS.includes(item.dataset.label) &&
                item.dataIndex === item.dataset.data.length - 1;
              if (isPumpForecast)
                return `${{currentFuel === 'diesel' ? 'Diesel' : 'Euro-95'}} price`;
              return item.raw?.sourceDate
                ? fmtDateLabel(item.raw.sourceDate)
                : item.label;
            }},
            label: ctx => {{
              if (ctx.dataset.label.startsWith('Brent'))
                return ` Brent: $${{ctx.parsed.y?.toFixed(2) ?? '—'}}/bbl`;
              const fuel = currentFuel === 'diesel' ? 'Diesel' : 'Euro-95';
              const isForecast = ctx.dataIndex === ctx.dataset.data.length - 1;
              if (!isForecast)
                return ` ${{ctx.dataset.label}} ${{fuel}} observed: €${{ctx.parsed.y?.toFixed(2) ?? '—'}}/L`;
              const current = ctx.dataset.data[ctx.dataset.data.length - 2].y;
              const forecastValue = ctx.parsed.y;
              const changeEuro = forecastValue - current;
              const changeLabel =
                `${{changeEuro >= 0 ? '+' : '-'}}${{Math.abs(changeEuro).toFixed(2)}}€`;
              return [
                ' Current',
                ` €${{current.toFixed(2)}}/L`,
                ' Forecast:',
                ` €${{forecastValue?.toFixed(2) ?? '—'}}/L (${{changeLabel}})`
              ];
            }}
          }}
        }}
      }},
      scales: {{
        x: {{
          type: 'linear',
          min: dateX(DATA.dates[0]),
          max: historyRightEdge(DATA.dates[0]),
          ticks: {{
            color: '#e2e8f0', maxTicksLimit: 18,
            font: {{ size: 13, weight: '600' }},
            callback: value => new Date(value).toLocaleDateString('en-GB', {{
              day:'numeric', month:'short', year:'2-digit', timeZone:'UTC'
            }})
          }},
          grid: {{ color: CHART_GRID }}
        }},
        y: {{
          position: 'right',
          min: 1.20,
          max: 2.80,
          ticks: {{
            stepSize: 0.20,
            color: '#e2e8f0',
            callback: v => `€${{v.toFixed(2)}}`,
            font: {{ size: 13, weight: '600' }}
          }},
          grid: {{ color: CHART_GRID }},
          title: {{ display: true, text: 'Pump price (€/L)', color: '#e2e8f0', font: {{ size: 13, weight: '600' }} }},
        }},
        y2: {{
          position: 'left',
          min: 60,
          max: 140,
          ticks: {{
            stepSize: 10,
            color: '#d4c98acc',
            callback: v => `$${{v.toFixed(0)}}`,
            font: {{ size: 13, weight: '600' }}
          }},
          grid: {{ display: false }},
          title: {{ display: true, text: 'Brent ($/bbl)', color: '#d4c98acc', font: {{ size: 13, weight: '600' }} }},
        }}
      }}
    }},
    plugins: [{{
      id: 'psychologicalPriceLine',
      beforeDatasetsDraw(chart) {{
        const scale = chart.scales.y;
        const y = scale.getPixelForValue(2.00);
        const {{ctx, chartArea}} = chart;
        if (y < chartArea.top || y > chartArea.bottom) return;

        ctx.save();
        ctx.setLineDash([6, 4]);
        ctx.strokeStyle = 'rgba(239, 68, 68, 0.8)';
        ctx.lineWidth = 1.25;
        ctx.beginPath();
        ctx.moveTo(chartArea.left, y);
        ctx.lineTo(chartArea.right, y);
        ctx.stroke();
        ctx.setLineDash([]);
        ctx.font = '600 10px Inter, sans-serif';
        ctx.textAlign = 'left';
        ctx.textBaseline = 'bottom';
        const label = '€2.00/L level';
        const width = ctx.measureText(label).width;
        ctx.fillStyle = 'rgba(2, 8, 23, 0.82)';
        ctx.fillRect(chartArea.left + 7, y - 17, width + 10, 15);
        ctx.fillStyle = '#ef4444';
        ctx.fillText(label, chartArea.left + 12, y - 4);
        ctx.restore();
      }}
    }}, {{
      id: 'forecastZone',
      beforeDatasetsDraw(chart) {{
        const firstPump = chart.data.datasets.findIndex(ds => CTRS.includes(ds.label));
        if (firstPump < 0) return;
        const dataset = chart.data.datasets[firstPump];
        const points = chart.getDatasetMeta(firstPump).data;
        const actual = points[dataset.data.length - 2];
        const forecast = points[dataset.data.length - 1];
        if (!actual || !forecast) return;

        const boundary = actual.x;
        const {{ctx, chartArea}} = chart;
        ctx.save();
        ctx.fillStyle = 'rgba(148, 163, 184, 0.055)';
        ctx.fillRect(boundary, chartArea.top, chartArea.right - boundary, chartArea.bottom - chartArea.top);
        ctx.setLineDash([4, 4]);
        ctx.strokeStyle = 'rgba(148, 163, 184, 0.65)';
        ctx.beginPath();
        ctx.moveTo(boundary, chartArea.top);
        ctx.lineTo(boundary, chartArea.bottom);
        ctx.stroke();
        ctx.setLineDash([]);
        ctx.fillStyle = '#94a3b8';
        ctx.font = '700 10px Inter, sans-serif';
        ctx.textAlign = 'left';
        ctx.textBaseline = 'top';
        ctx.fillText('MODEL FORECAST', boundary + 8, chartArea.top + 8);
        ctx.restore();
      }}
    }}, {{
      id: 'countryFocus',
      beforeDatasetDraw(chart, args) {{
        chart.ctx.save();
        const focus = chart.$hoverCountry;
        const label = chart.data.datasets[args.index]?.label;
        if (focus != null && CTRS.includes(label) && args.index !== focus)
          chart.ctx.globalAlpha = 0.18;
      }},
      afterDatasetDraw(chart) {{
        chart.ctx.restore();
      }}
    }}, {{
      id: 'latestBrentLabel',
      afterDatasetsDraw(chart) {{
        const datasetIndex = chart.data.datasets.findIndex(ds => ds.label.startsWith('Brent'));
        if (datasetIndex < 0) return;
        const dataset = chart.data.datasets[datasetIndex];
        const lastIndex = dataset.data.length - 1;
        const value = dataset.data[lastIndex]?.y;
        const point = chart.getDatasetMeta(datasetIndex).data[lastIndex];
        if (value == null || !point) return;

        const observationDate = new Date(
          (DATA.brent_latest?.date ?? DATA.dates[DATA.dates.length - 1]) + 'T12:00:00Z'
        );
        const shiftedDate = observationDate.toLocaleDateString('en-GB', {{
          day: 'numeric', month: 'short', year: '2-digit', timeZone: 'UTC'
        }});
        const move = latestBrentMove();
        const moveLabel = move == null ? '' : ` (${{move >= 0 ? '+' : '-'}}$${{Math.abs(move).toFixed(1)}})`;
        const titleLabel = 'Brent';
        const valueLabel = `$${{value.toFixed(0)}}/bbl${{moveLabel}}`;
        const ctx = chart.ctx;
        ctx.save();
        ctx.font = '600 11px Inter, sans-serif';
        ctx.textAlign = 'left';
        ctx.textBaseline = 'top';
        const width = Math.max(
          ctx.measureText(titleLabel).width,
          ctx.measureText(valueLabel).width,
          ctx.measureText(shiftedDate).width
        );
        const x = point.x + 9;
        const y = point.y - 25;
        ctx.fillStyle = 'rgba(2, 8, 23, 0.82)';
        ctx.fillRect(x - 5, y - 4, width + 10, 46);
        ctx.fillStyle = '#cbd5e1';
        ctx.fillText(titleLabel, x, y);
        ctx.fillStyle = BRENT_COLOR;
        ctx.fillText(valueLabel, x, y + 14);
        ctx.fillStyle = '#cbd5e1';
        ctx.fillText(shiftedDate, x, y + 28);
        ctx.restore();
      }}
    }}]
  }});
  buildHistoryLegend();
}}

function updateHistChart() {{
  const total = DATA.dates.length;
  const start = currentRange === 0 ? 0 : Math.max(0, total - currentRange);
  const startDate = DATA.dates[start];
  histChart.data.datasets.forEach((ds, i) => {{
    if (ds.label.startsWith('Brent')) {{
      ds.data = displayedBrentSeries(startDate);
    }} else if (ds.label.startsWith('Pump forecast')) {{
      ds.data = [];
    }} else {{
      const c = CTRS[i];
      ds.data = pumpHistoryWithForecast(c).filter(
        point => point.sourceDate >= startDate
      );
    }}
  }});
  histChart.options.scales.x.min = dateX(startDate);
  histChart.options.scales.x.max = historyRightEdge(startDate);
  const pumpBounds = sharedPumpAxisBounds(start);
  if (pumpBounds) {{
    histChart.options.scales.y.min = pumpBounds.min;
    histChart.options.scales.y.max = pumpBounds.max;
    histChart.options.scales.y2.min = pumpBounds.brentMin;
    histChart.options.scales.y2.max = pumpBounds.brentMax;
  }}
  histChart.update();
}}

function switchFuel(fuel) {{
  currentFuel = fuel;
  $('btn95').classList.toggle('active', fuel === 'euro95');
  $('btnD' ).classList.toggle('active', fuel === 'diesel');
  updateRefuelCallout();
  updateHistChart();
}}

function setRange(n) {{
  currentRange = n;
  $('btnYTD').classList.toggle('active', n === DATA.ytd_weeks);
  $('btn1Y' ).classList.toggle('active', n === 52);
  $('btn3Y' ).classList.toggle('active', n === 156);
  $('btn5Y' ).classList.toggle('active', n === 260);
  $('btnAll').classList.toggle('active', n === 0);
  updateHistChart();
}}

// ---- PRICE CHART ---------------------------------------------------------
function buildPriceChart() {{
  const last      = DATA.dates.length - 1;
  const dieselData = CTRS.map(c => +((DATA.countries[c].diesel[last] || 0) / 1000).toFixed(4));
  const euro95Data = CTRS.map(c => +((DATA.countries[c].euro95[last] || 0) / 1000).toFixed(4));

  new Chart($('priceChart').getContext('2d'), {{
    type: 'bar',
    data: {{
      labels: CTRS,
      datasets: [
        {{
          label: 'Diesel',
          data: dieselData,
          backgroundColor: CTRS.map(c => COLORS[c] + 'cc'),
          borderRadius: 4,
        }},
        {{
          label: 'Euro-95',
          data: euro95Data,
          backgroundColor: CTRS.map(c => COLORS[c] + '55'),
          borderRadius: 4,
        }},
      ],
    }},
    options: {{
      responsive: true, maintainAspectRatio: false, devicePixelRatio: window.devicePixelRatio,
      layout: {{ padding: {{ top: 34 }} }},
      plugins: {{
        legend: {{ display: false }},
        tooltip: {{
          backgroundColor: TOOLTIP_BG, borderColor: 'rgba(148,163,184,.2)', borderWidth: 1,
          titleColor: '#94a3b8', bodyColor: '#e2e8f0',
          callbacks: {{
            title: items => items.length ? `${{items[0].label}} · ${{items[0].dataset.label}}` : '',
            label: ctx => {{
              const fuelKey = ctx.datasetIndex === 0 ? 'diesel' : 'euro95';
              const move = latestBrentMove();
              const coefficient = forecastCoefficientForFuel(fuelKey, move);
              const forecast = ctx.raw + coefficient * move / 10;
              const changeCents = (forecast - ctx.raw) * 100;
              const roundedCents = Math.round(changeCents);
              const sign = roundedCents >= 0 ? '+' : '';
              return [
                ' Current',
                ` €${{ctx.raw.toFixed(2)}}/L`,
                ' Next Week Forecast',
                ` €${{forecast.toFixed(2)}}/L (${{sign}}${{roundedCents}} cents)`
              ];
            }}
          }}
        }},
      }},
      scales: {{
        x: {{ ticks: {{ color: '#e2e8f0', font: {{ size: 12, weight: 700 }} }}, grid: {{ color: CHART_GRID }} }},
        y: {{
          min: 1.30,
          grace: '10%',
          ticks: {{ color: '#e2e8f0', callback: v => `€${{v.toFixed(2)}}`, font: {{ size: 13, weight: '600' }} }},
          grid: {{ color: CHART_GRID }},
          title: {{ display: true, text: '€/L', color: '#e2e8f0', font: {{ size: 13, weight: '600' }} }},
        }},
      }},
    }},
    plugins: [{{
      id: 'priceLabels',
      afterDatasetsDraw(chart) {{
        const ctx2 = chart.ctx;
        const fuelLabels = ['Diesel', '95'];
        const brentMove = latestBrentMove();
        const direction = brentMove == null ? null : (brentMove >= 0 ? '▲' : '▼');
        chart.data.datasets.forEach((ds, i) => {{
          chart.getDatasetMeta(i).data.forEach((bar, j) => {{
            const val = ds.data[j];
            if (val == null) return;
            ctx2.save();
            ctx2.textAlign = 'center';
            // fuel type
            ctx2.font = '10px DM Sans, sans-serif';
            ctx2.fillStyle = '#cbd5e1';
            ctx2.textBaseline = 'bottom';
            ctx2.fillText(fuelLabels[i], bar.x, chart.chartArea.top - 18);
            // value
            ctx2.font = 'bold 11px DM Sans, sans-serif';
            ctx2.fillStyle = '#f8fafc';
            const valueText = `€${{val.toFixed(2)}}`;
            ctx2.fillText(valueText, bar.x, chart.chartArea.top - 4);

            // Prediction direction: a large two-part arrow above the bar.
            if (direction) {{
              const color = Array.isArray(ds.backgroundColor)
                ? ds.backgroundColor[j] : ds.backgroundColor;
              const x = bar.x;
              ctx2.fillStyle = color;
              ctx2.beginPath();
              if (brentMove >= 0) {{
                ctx2.moveTo(x, bar.y - 22);
                ctx2.lineTo(x - 8, bar.y - 11);
                ctx2.lineTo(x + 8, bar.y - 11);
                ctx2.closePath();
                ctx2.fill();
                ctx2.fillRect(x - 7, bar.y - 7, 14, 4);
              }} else {{
                ctx2.fillRect(x - 7, bar.y - 22, 14, 4);
                ctx2.moveTo(x - 8, bar.y - 14);
                ctx2.lineTo(x + 8, bar.y - 14);
                ctx2.lineTo(x, bar.y - 3);
                ctx2.closePath();
                ctx2.fill();
              }}
            }}
            ctx2.restore();
          }});
        }});
      }}
    }}],
  }});
}}

// ---- YTD -----------------------------------------------------------------
function buildYTD() {{
  const yr = DATA.dates[DATA.dates.length-1].slice(0,4);
  $('ytd-year').textContent = yr;
  $('ytd-tab-label').textContent = `${{yr}} Variation`;

  // Cards
  const wrap = $('ytd-cards');
  CTRS.forEach(c => {{
    const y = DATA.ytd[c];
    const el = document.createElement('div');
    el.className = 'ytd-card';
    el.style.borderColor = COLORS[c] + '40';
    const cls95 = (y.euro95_ytd ?? 0) < 0 ? 'dn' : 'up';
    const clsD  = (y.diesel_ytd  ?? 0) < 0 ? 'dn' : 'up';
    function fmtLine(abs, pct) {{
      if (abs == null && pct == null) return '—';
      const absStr = abs != null ? `${{abs>=0?'+':''}}${{abs.toFixed(2)}} €` : '';
      const pctStr = pct != null ? `(${{pct>=0?'+':''}}${{pct.toFixed(1)}}%)` : '';
      return absStr && pctStr ? `${{absStr}} ${{pctStr}}` : absStr || pctStr;
    }}
    el.innerHTML = `
      <div class="ytd-ctr" style="color:${{COLORS[c]}}">${{c}}</div>
      <div class="ytd-sub">Diesel</div>
      <div class="ytd-val ${{clsD}} mono">${{fmtLine(y.diesel_abs, y.diesel_ytd)}}</div>
      <div class="ytd-sub" style="margin-top:6px">Euro-95</div>
      <div class="ytd-val2 ${{cls95}} mono">${{fmtLine(y.euro95_abs, y.euro95_ytd)}}</div>`;
    wrap.appendChild(el);
  }});

  // Brent card
  if (DATA.brent_ytd != null) {{
    const clsB = DATA.brent_ytd < 0 ? 'dn' : 'up';
    const el = document.createElement('div');
    el.className = 'ytd-card';
    el.style.borderColor = '#94a3b840';
    el.innerHTML = `
      <div class="ytd-ctr" style="color:#94a3b8">Brent</div>
      <div class="ytd-sub">$/barrel</div>
      <div class="ytd-val ${{clsB}} mono">${{(DATA.brent_ytd>0?'+':'')+DATA.brent_ytd.toFixed(2)+'%'}}</div>`;
    wrap.appendChild(el);
  }}

  // ---- Expected vs Actual insight box ------------------------------------
  if (DATA.brent_abs != null) {{
    const brentMove   = DATA.brent_abs;   // $/bbl YTD
    // Avg sensitivity slope (eurocents/L per $10 Brent) across diesel + 95, 4W lag1W
    const sensD   = DATA.sensitivity.diesel.research['4_1'].up;
    const sens95  = DATA.sensitivity.euro95.research['4_1'].up;
    const avgSens = (sensD + sens95) / 2;                    // eurocents/L per $10
    const expectedCents = (brentMove / 10) * avgSens;        // eurocents/L
    const expectedEuro  = (expectedCents / 100).toFixed(2);  // €/L

    // Actual EU avg diesel change (countries excl. EU aggregate)
    const ctrsNoEU = CTRS.filter(c => c !== 'EU');
    const actualDiesel = ctrsNoEU
      .map(c => DATA.ytd[c].diesel_abs)
      .filter(v => v != null);
    const avgActual = actualDiesel.reduce((s,v) => s+v, 0) / actualDiesel.length;
    const overshootPct = Math.round((avgActual / (expectedCents/100) - 1) * 100);
    const sign = brentMove >= 0 ? '+' : '';

    $('ytd-insight').innerHTML = `
      <div style="margin:12px 0;padding:14px 20px;background:rgba(245,158,11,0.08);
                  border:1px solid rgba(245,158,11,0.3);border-radius:10px;
                  font-size:13px;line-height:1.7;color:#cbd5e1;">
        <span style="color:#f59e0b;font-weight:700;">Model check · Brent YTD ${{sign}}$${{brentMove.toFixed(2)}}/bbl</span><br>
        At our pass-through sensitivity of
        <strong style="color:#f1f5f9;">+${{avgSens.toFixed(1)}}¢/L per $10 Brent</strong>,
        a ${{sign}}$${{brentMove.toFixed(2)}} move implies an expected pump price change of
        <strong style="color:#f1f5f9;">~${{brentMove >= 0 ? '+' : ''}}${{expectedEuro}}€/L</strong>.<br>
        Actual EU diesel average: <strong style="color:#f59e0b;">${{avgActual >= 0 ? '+' : ''}}${{avgActual.toFixed(2)}}€/L</strong>
        — <strong style="color:#f87171;">${{Math.abs(overshootPct)}}% ${{overshootPct > 0 ? 'above' : 'below'}} model</strong>.
      </div>`;
  }}

  // Bar chart
  const ytdData95   = CTRS.map(c => DATA.ytd[c].euro95_ytd);
  const ytdDataD    = CTRS.map(c => DATA.ytd[c].diesel_ytd);
  const chartLabels = [...CTRS, 'Brent'];

  ytdChart = new Chart($('ytdChart').getContext('2d'), {{
    type: 'bar',
    data: {{
      labels: chartLabels,
      datasets: [
        {{
          label: 'Diesel YTD%',
          data: [...ytdDataD, null],
          backgroundColor: [...CTRS.map(c => COLORS[c] + 'cc'), 'transparent'],
          borderRadius: 4,
          yAxisID: 'y',
        }},
        {{
          label: 'Euro-95 YTD%',
          data: [...ytdData95, null],
          backgroundColor: [...CTRS.map(c => COLORS[c] + '55'), 'transparent'],
          borderRadius: 4,
          yAxisID: 'y',
        }},
        {{
          label: 'Brent YTD%',
          data: [...CTRS.map(() => null), DATA.brent_ytd],
          backgroundColor: '#94a3b880',
          borderColor: '#94a3b8',
          borderWidth: 1,
          borderRadius: 4,
          yAxisID: 'y2',
          order: 0,
        }},
      ]
    }},
    options: {{
      responsive: true, maintainAspectRatio: false, devicePixelRatio: window.devicePixelRatio,
      plugins: {{
        legend: {{ labels: {{ color: '#94a3b8', font: {{ size: 11 }} }} }},
        tooltip: {{
          backgroundColor: TOOLTIP_BG, borderColor: 'rgba(148,163,184,.2)', borderWidth: 1,
          titleColor: '#94a3b8', bodyColor: '#e2e8f0',
          callbacks: {{ label: ctx => ` ${{ctx.dataset.label}}: ${{ctx.raw?.toFixed(2)}}%` }}
        }}
      }},
      scales: {{
        x: {{ ticks: {{ color: '#e2e8f0', font: {{ size: 12, weight: 700 }} }}, grid: {{ color: CHART_GRID }} }},
        y: {{
          ticks: {{ color: '#e2e8f0', callback: v => v+'%', font: {{ size: 13, weight: '600' }} }},
          grid: {{ color: CHART_GRID }},
          title: {{ display: true, text: 'Pump price YTD (%)', color: '#e2e8f0', font: {{ size: 13, weight: '600' }} }},
        }},
        y2: {{
          type: 'linear',
          position: 'right',
          ticks: {{ color: '#e2e8f0', callback: v => v+'%', font: {{ size: 13, weight: '600' }} }},
          grid: {{ display: false }},
          title: {{ display: true, text: 'Brent YTD (%)', color: '#e2e8f0', font: {{ size: 13, weight: '600' }} }},
        }}
      }}
    }},
    plugins: [{{
      id: 'ytdLabels',
      afterDatasetsDraw(chart) {{
        const ctx2 = chart.ctx;
        chart.data.datasets.forEach((ds, i) => {{
          chart.getDatasetMeta(i).data.forEach((bar, j) => {{
            const val = ds.data[j];
            if (val == null) return;
            const text = (val >= 0 ? '+' : '') + val.toFixed(1) + '%';
            ctx2.save();
            ctx2.fillStyle = '#e2e8f0';
            ctx2.font = 'bold 10px DM Sans, sans-serif';
            ctx2.textAlign = 'center';
            ctx2.textBaseline = val >= 0 ? 'bottom' : 'top';
            ctx2.fillText(text, bar.x, val >= 0 ? bar.y - 3 : bar.y + 3);
            ctx2.restore();
          }});
        }});
      }}
    }}, {{
      id: 'ytdDivider',
      afterDraw(chart) {{
        const xScale = chart.scales.x;
        const sepIdx = CTRS.length - 1;  // between PT (last country) and Brent
        const x1 = xScale.getPixelForValue(sepIdx);
        const x2 = xScale.getPixelForValue(sepIdx + 1);
        const midX = (x1 + x2) / 2;
        const {{ top, bottom }} = chart.chartArea;
        const ctx2 = chart.ctx;
        ctx2.save();
        ctx2.strokeStyle = '#94a3b8';
        ctx2.lineWidth = 1;
        ctx2.setLineDash([5, 4]);
        ctx2.beginPath();
        ctx2.moveTo(midX, top);
        ctx2.lineTo(midX, bottom);
        ctx2.stroke();
        ctx2.setLineDash([]);
        ctx2.restore();
      }}
    }}]
  }});

  // ---- Absolute change chart (€/L pump · $/bbl Brent) --------------------
  const absD   = CTRS.map(c => DATA.ytd[c].diesel_abs  != null ? +DATA.ytd[c].diesel_abs.toFixed(4)  : null);
  const abs95  = CTRS.map(c => DATA.ytd[c].euro95_abs  != null ? +DATA.ytd[c].euro95_abs.toFixed(4)  : null);
  new Chart($('ytdAbsChart').getContext('2d'), {{
    type: 'bar',
    data: {{
      labels: [...CTRS, 'Brent'],
      datasets: [
        {{
          label: 'Diesel',
          data: [...absD, null],
          backgroundColor: [...CTRS.map(c => COLORS[c] + 'cc'), 'transparent'],
          borderRadius: 4, yAxisID: 'y',
        }},
        {{
          label: 'Euro-95',
          data: [...abs95, null],
          backgroundColor: [...CTRS.map(c => COLORS[c] + '55'), 'transparent'],
          borderRadius: 4, yAxisID: 'y',
        }},
        {{
          label: 'Brent',
          data: [...CTRS.map(() => null), DATA.brent_abs],
          backgroundColor: '#94a3b880', borderColor: '#94a3b8', borderWidth: 1,
          borderRadius: 4, yAxisID: 'y2', order: 0,
        }},
      ]
    }},
    options: {{
      responsive: true, maintainAspectRatio: false, devicePixelRatio: window.devicePixelRatio,
      plugins: {{
        legend: {{ labels: {{ color: '#94a3b8', font: {{ size: 11 }} }} }},
        tooltip: {{
          backgroundColor: TOOLTIP_BG, borderColor: 'rgba(148,163,184,.2)', borderWidth: 1,
          titleColor: '#94a3b8', bodyColor: '#e2e8f0',
          callbacks: {{
            label: ctx => ctx.datasetIndex < 2
              ? ` ${{ctx.dataset.label}}: ${{ctx.raw >= 0 ? '+' : ''}}${{ctx.raw?.toFixed(2)}}€/L`
              : ` Brent: ${{ctx.raw >= 0 ? '+' : ''}}$${{ctx.raw?.toFixed(2)}}/bbl`
          }}
        }}
      }},
      scales: {{
        x: {{ ticks: {{ color: '#e2e8f0', font: {{ size: 12, weight: 700 }} }}, grid: {{ color: CHART_GRID }} }},
        y: {{
          ticks: {{ color: '#e2e8f0', callback: v => (v >= 0 ? '+' : '') + v.toFixed(2) + '€', font: {{ size: 13, weight: '600' }} }},
          grid: {{ color: CHART_GRID }},
          title: {{ display: true, text: 'Pump price change (€/L)', color: '#e2e8f0', font: {{ size: 13, weight: '600' }} }},
        }},
        y2: {{
          type: 'linear', position: 'right',
          ticks: {{ color: '#e2e8f0', callback: v => (v >= 0 ? '+$' : '-$') + Math.abs(v), font: {{ size: 13, weight: '600' }} }},
          grid: {{ display: false }},
          title: {{ display: true, text: 'Brent change ($/bbl)', color: '#e2e8f0', font: {{ size: 13, weight: '600' }} }},
        }}
      }}
    }},
    plugins: [{{
      id: 'absLabels',
      afterDatasetsDraw(chart) {{
        const ctx2 = chart.ctx;
        chart.data.datasets.forEach((ds, i) => {{
          chart.getDatasetMeta(i).data.forEach((bar, j) => {{
            const val = ds.data[j];
            if (val == null) return;
            const isBrent = i === 2;
            const text = isBrent
              ? (val >= 0 ? '+$' : '-$') + Math.abs(val).toFixed(2)
              : (val >= 0 ? '+' : '') + val.toFixed(2) + '€';
            ctx2.save();
            ctx2.fillStyle = '#e2e8f0';
            ctx2.font = 'bold 10px DM Sans, sans-serif';
            ctx2.textAlign = 'center';
            ctx2.textBaseline = val >= 0 ? 'bottom' : 'top';
            ctx2.fillText(text, bar.x, val >= 0 ? bar.y - 3 : bar.y + 3);
            ctx2.restore();
          }});
        }});
      }}
    }}, {{
      id: 'absDivider',
      afterDraw(chart) {{
        const xScale = chart.scales.x;
        const x1 = xScale.getPixelForValue(CTRS.length - 1);
        const x2 = xScale.getPixelForValue(CTRS.length);
        const midX = (x1 + x2) / 2;
        const {{ top, bottom }} = chart.chartArea;
        const ctx2 = chart.ctx;
        ctx2.save();
        ctx2.strokeStyle = '#94a3b8';
        ctx2.lineWidth = 1;
        ctx2.setLineDash([5, 4]);
        ctx2.beginPath();
        ctx2.moveTo(midX, top);
        ctx2.lineTo(midX, bottom);
        ctx2.stroke();
        ctx2.setLineDash([]);
        ctx2.restore();
      }}
    }}]
  }});
}}

// ---- TAX CHARTS ----------------------------------------------------------
function buildTaxCharts() {{
  const last = DATA.dates.length - 1;
  const ctrs = CTRS.filter(c => c !== 'EU');

  // Diagonal-stripe hatch pattern for the Tax & Duties segment
  function makeHatch(color) {{
    const sz = 8;
    const cv = document.createElement('canvas');
    cv.width = sz; cv.height = sz;
    const cx = cv.getContext('2d');
    cx.fillStyle = color + 'aa';
    cx.fillRect(0, 0, sz, sz);
    cx.strokeStyle = color;
    cx.lineWidth = 2;
    cx.beginPath();
    cx.moveTo(-1, sz + 1); cx.lineTo(sz + 1, -1);
    cx.moveTo(-1 - sz, sz + 1); cx.lineTo(sz + 1 - sz, -1);
    cx.moveTo(-1 + sz, sz + 1); cx.lineTo(sz + 1 + sz, -1);
    cx.stroke();
    return cx.createPattern(cv, 'repeat');
  }}

  function taxDatasets(fuelKey, fuelKeyNt) {{
    return [
      {{
        label: 'Pre-tax',
        data: ctrs.map(c => {{
          const v = DATA.countries[c][fuelKeyNt][last];
          return v != null ? +(v/1000).toFixed(4) : null;
        }}),
        backgroundColor: ctrs.map(c => COLORS[c]),
        borderRadius: [0,0,0,0],
        stack: 'a',
      }},
      {{
        label: 'Tax & Duties',
        data: ctrs.map(c => {{
          const gross = DATA.countries[c][fuelKey][last];
          const net   = DATA.countries[c][fuelKeyNt][last];
          if (gross == null || net == null) return null;
          return +((gross - net)/1000).toFixed(4);
        }}),
        backgroundColor: ctrs.map(c => COLORS[c] + '55'),
        borderRadius: [4,4,0,0],
        stack: 'a',
      }},
    ];
  }}

  // Plugin: value labels inside each stacked segment
  const stackLabels = {{
    id: 'stackLabels',
    afterDatasetsDraw(chart) {{
      const ctx2 = chart.ctx;
      chart.data.datasets.forEach((ds, i) => {{
        chart.getDatasetMeta(i).data.forEach((bar, j) => {{
          const val = ds.data[j];
          if (val == null || val < 0.05) return;
          const segH = Math.abs(bar.base - bar.y);
          if (segH < 14) return;
          const midY = (bar.y + bar.base) / 2;
          ctx2.save();
          ctx2.fillStyle = '#fff';
          ctx2.font = 'bold 10px DM Sans, sans-serif';
          ctx2.textAlign = 'center';
          ctx2.textBaseline = 'middle';
          ctx2.fillText(`€${{val.toFixed(2)}}`, bar.x, midY);
          ctx2.restore();
        }});
      }});
    }}
  }};

  const baseOpts = {{
    responsive: true, maintainAspectRatio: false, devicePixelRatio: window.devicePixelRatio,
    plugins: {{
      legend: {{ labels: {{ color:'#94a3b8', font:{{ size:10 }} }} }},
      tooltip: {{
        backgroundColor:TOOLTIP_BG, borderColor:'rgba(148,163,184,.2)', borderWidth:1,
        titleColor:'#94a3b8', bodyColor:'#e2e8f0',
        callbacks: {{ label: ctx => ` ${{ctx.dataset.label}}: €${{ctx.raw?.toFixed(3) ?? '—'}}/L` }}
      }}
    }},
    scales: {{
      x: {{ stacked:true, ticks:{{ color:'#e2e8f0', font:{{ size:11 }} }}, grid:{{ color:CHART_GRID }} }},
      y: {{ stacked:true,
            ticks:{{ color:'#e2e8f0', callback: v=>`€${{v.toFixed(2)}}`, font:{{ size:10 }} }},
            grid:{{ color:CHART_GRID }} }}
    }}
  }};

  taxDChart = new Chart($('taxDChart').getContext('2d'), {{
    type:'bar',
    data: {{ labels: ctrs, datasets: taxDatasets('diesel','diesel_notax') }},
    options: baseOpts,
    plugins: [stackLabels],
  }});
  tax95Chart = new Chart($('tax95Chart').getContext('2d'), {{
    type:'bar',
    data: {{ labels: ctrs, datasets: taxDatasets('euro95','euro95_notax') }},
    options: baseOpts,
    plugins: [stackLabels],
  }});

  // Tables under each chart
  function buildTaxTable(tableId, fuelKey, fuelKeyNt) {{
    const wrap = $(tableId);
    const tbl = document.createElement('table');
    tbl.style.cssText = 'width:100%;border-collapse:collapse;font-size:11px;margin-top:8px;';
    tbl.innerHTML = `<thead><tr>
      <th style="text-align:left;padding:4px 8px;color:#94a3b8;font-weight:600;border-bottom:1px solid #1e293b;">Country</th>
      <th style="text-align:right;padding:4px 8px;color:#94a3b8;font-weight:600;border-bottom:1px solid #1e293b;">Pre-tax</th>
      <th style="text-align:right;padding:4px 8px;color:#94a3b8;font-weight:600;border-bottom:1px solid #1e293b;">Tax & Duties</th>
      <th style="text-align:right;padding:4px 8px;color:#94a3b8;font-weight:600;border-bottom:1px solid #1e293b;">Total</th>
      <th style="text-align:right;padding:4px 8px;color:#94a3b8;font-weight:600;border-bottom:1px solid #1e293b;">Tax Rate</th>
    </tr></thead><tbody></tbody>`;
    const tbody = tbl.querySelector('tbody');
    ctrs.forEach(c => {{
      const gross  = DATA.countries[c][fuelKey][last];
      const net    = DATA.countries[c][fuelKeyNt][last];
      const pretax = net   != null ? `€${{(net/1000).toFixed(3)}}` : '—';
      const tax    = (gross && net) ? `€${{((gross-net)/1000).toFixed(3)}}` : '—';
      const total  = gross != null ? `€${{(gross/1000).toFixed(3)}}` : '—';
      const rate   = (gross && net) ? Math.round((gross-net)/gross*100)+'%' : '—';
      const tr = document.createElement('tr');
      tr.innerHTML = `
        <td style="padding:4px 8px;color:${{COLORS[c]}};font-weight:700;">${{c}}</td>
        <td class="mono" style="text-align:right;padding:4px 8px;color:#e2e8f0;">${{pretax}}</td>
        <td class="mono" style="text-align:right;padding:4px 8px;color:#e2e8f0;">${{tax}}</td>
        <td class="mono" style="text-align:right;padding:4px 8px;color:#f1f5f9;font-weight:700;">${{total}}</td>
        <td class="mono" style="text-align:right;padding:4px 8px;color:#94a3b8;">${{rate}}</td>`;
      tbody.appendChild(tr);
    }});
    wrap.appendChild(tbl);
  }}
  buildTaxTable('taxDTable', 'diesel', 'diesel_notax');
  buildTaxTable('tax95Table', 'euro95', 'euro95_notax');

  // Tax rate horizontal bars
  const barWrap = $('tax-bars');
  const sorted = [...ctrs].sort((a,b) => DATA.countries[b].tax_rate - DATA.countries[a].tax_rate);
  sorted.forEach(c => {{
    const pct = Math.round(DATA.countries[c].tax_rate * 100);
    const last_95 = DATA.countries[c].euro95[DATA.dates.length-1];
    const last_nt = DATA.countries[c].euro95_notax[DATA.dates.length-1];
    const taxAmt  = (last_95 && last_nt) ? ((last_95 - last_nt)/1000).toFixed(3) : '—';
    const el = document.createElement('div');
    el.className = 'tax-bar-row';
    el.innerHTML = `
      <div class="tax-bar-label" style="color:${{COLORS[c]}}">${{c}}</div>
      <div class="tax-bar-track">
        <div class="tax-bar-fill" style="width:${{pct}}%;background:linear-gradient(90deg,${{COLORS[c]}}60,${{COLORS[c]}})">
          <span class="tax-bar-pct mono">${{pct}}%</span>
        </div>
      </div>
      <div class="tax-bar-aside mono">€${{taxAmt}}/L tax</div>`;
    barWrap.appendChild(el);
  }});
}}

// ---- CONSUMPTION ---------------------------------------------------------
function switchCons(mode) {{
  currentCons = mode;
  $('btnAbsolute').classList.toggle('active', mode === 'absolute');
  $('btnMix'     ).classList.toggle('active', mode === 'mix');
  $('cons-absolute').style.display = mode === 'absolute' ? '' : 'none';
  $('cons-mix'     ).style.display = mode === 'mix'      ? '' : 'none';
}}

// Shared legend options for both cons charts (no generateLabels override — let Chart.js do it correctly)
function consLegendOpts(tooltipFmt) {{
  return {{
    responsive: true, maintainAspectRatio: false, devicePixelRatio: window.devicePixelRatio,
    plugins: {{
      legend: {{
        position: 'top',
        labels: {{
          color: '#94a3b8',
          font: {{ size: 11 }},
          boxWidth: 14,
          boxHeight: 14,
          padding: 16,
          usePointStyle: false,
        }}
      }},
      tooltip: {{
        backgroundColor: TOOLTIP_BG, borderColor: 'rgba(148,163,184,.2)', borderWidth: 1,
        titleColor: '#94a3b8', bodyColor: '#e2e8f0',
        callbacks: {{ label: tooltipFmt }}
      }}
    }},
    scales: {{
      x: {{ stacked: true, ticks: {{ color:'#e2e8f0', font:{{ size:13, weight:'700' }} }}, grid:{{ color:CHART_GRID }} }},
      y: {{ stacked: true, ticks: {{ color:'#e2e8f0', font:{{ size:10 }} }},               grid:{{ color:CHART_GRID }} }}
    }}
  }};
}}

function buildConsumption() {{
  const ctrs = CTRS.filter(c => c !== 'EU');
  $('cons-year').textContent = DATA.latest_year;

  // ---- Absolute volumes chart -------------------------------------------
  const absOpts = consLegendOpts(ctx => ` ${{ctx.dataset.label}}: ${{ctx.raw != null ? ctx.raw.toLocaleString('en',{{maximumFractionDigits:0}}) : '—'}} kt`);
  absOpts.scales.y.ticks.callback = v => v >= 1000 ? (v/1000).toFixed(0)+'k' : v;

  consAbsChart = new Chart($('consAbsChart').getContext('2d'), {{
    type: 'bar',
    data: {{
      labels: ctrs,
      datasets: FUEL_TYPES.map(f => ({{
        label: FUEL_DISPLAY[f],
        data: ctrs.map(c => DATA.consumption[c][f] || 0),
        backgroundColor: FUEL_COLORS[f],
        stack: 'a',
      }}))
    }},
    options: absOpts,
  }});

  // ---- Percentage mix chart ---------------------------------------------
  const pctData = ctrs.map(c => {{
    const vals  = DATA.consumption[c];
    const total = Object.values(vals).reduce((a,b)=>a+b, 0);
    const row   = {{}};
    FUEL_TYPES.forEach(f => {{ row[f] = total > 0 ? +((vals[f]||0)/total*100).toFixed(1) : 0; }});
    return row;
  }});

  const mixOpts = consLegendOpts(ctx => ` ${{ctx.dataset.label}}: ${{ctx.raw?.toFixed(1) ?? '—'}}%`);
  mixOpts.scales.y.min = 0;
  mixOpts.scales.y.max = 100;
  mixOpts.scales.y.ticks.callback = v => v + '%';

  consMixChart = new Chart($('consMixChart').getContext('2d'), {{
    type: 'bar',
    data: {{
      labels: ctrs,
      datasets: FUEL_TYPES.map(f => ({{
        label: FUEL_DISPLAY[f],
        data: pctData.map(r => r[f]),
        backgroundColor: FUEL_COLORS[f],
        stack: 'a',
      }}))
    }},
    options: mixOpts,
  }});

  // ---- Absolute table ---------------------------------------------------
  const thead = $('cons-head');
  ['Country', ...FUEL_TYPES, 'Total'].forEach(h => {{
    const th = document.createElement('th');
    th.textContent = FUEL_DISPLAY[h] || h;
    th.style.color = FUEL_COLORS[h] || '#94a3b8';
    if (h === 'Country') th.style.textAlign = 'left';
    thead.appendChild(th);
  }});
  const tbody = $('cons-body');
  ctrs.forEach(c => {{
    const v     = DATA.consumption[c];
    const total = Object.values(v).reduce((a,b) => a+b, 0);
    const tr    = document.createElement('tr');
    tr.innerHTML = `
      <td><span class="dot" style="background:${{COLORS[c]}}"></span>
          <span style="color:${{COLORS[c]}};font-weight:700">${{c}}</span></td>
      ${{FUEL_TYPES.map(f => `<td class="mono" style="color:#94a3b8">${{fmtVal(v[f]||0, 0)}}</td>`).join('')}}
      <td class="mono" style="font-weight:700;color:#f1f5f9">${{fmtVal(total, 0)}}</td>`;
    tbody.appendChild(tr);
  }});
}}
// ---- SENSITIVITY ---------------------------------------------------------
function buildSensitivity() {{
  // All OLS slopes are pre-computed in Python; JS only handles rendering.
  function makeBarChart(canvasId, slopeId, fuelKey) {{
    const sens = DATA.sensitivity[fuelKey].per_country;
    const upData   = CTRS.map(c => sens[c].up);
    const downData = CTRS.map(c => sens[c].down);
    const avgUp   = ((upData.reduce((s, v) => s + v, 0) / upData.length) / 100).toFixed(2);
    const avgDown = ((downData.reduce((s, v) => s + v, 0) / downData.length) / 100).toFixed(2);
    $(slopeId).textContent = `avg ▲ +${{avgUp}}€/L  ▼ +${{avgDown}}€/L  per $10 Brent`;

    return new Chart($(canvasId).getContext('2d'), {{
      type: 'bar',
      data: {{
        labels: CTRS,
        datasets: [
          {{
            label: 'Brent ▲ (rising weeks)',
            data: upData,
            backgroundColor: CTRS.map(c => COLORS[c] + 'cc'),
            borderRadius: 4,
          }},
          {{
            label: 'Brent ▼ (falling weeks)',
            data: downData,
            backgroundColor: CTRS.map(c => COLORS[c] + '55'),
            borderRadius: 4,
          }},
        ],
      }},
      options: {{
        responsive: true, maintainAspectRatio: false, devicePixelRatio: window.devicePixelRatio,
        plugins: {{
          legend: {{ labels: {{ color: '#94a3b8', font: {{ size: 11 }} }} }},
          tooltip: {{
            backgroundColor: TOOLTIP_BG, borderColor: 'rgba(148,163,184,.2)', borderWidth: 1,
            titleColor: '#94a3b8', bodyColor: '#e2e8f0',
            callbacks: {{ label: ctx => ` ${{ctx.dataset.label}}: +${{(ctx.raw/100)?.toFixed(2)}}€/L per $10 Brent` }}
          }}
        }},
        scales: {{
          x: {{ ticks: {{ color: '#e2e8f0', font: {{ size: 12, weight: 700 }} }}, grid: {{ color: CHART_GRID }} }},
          y: {{
            min: 0,
            ticks: {{ color: '#e2e8f0', callback: v => '+' + (v/100).toFixed(2) + '€', font: {{ size: 13, weight: '600' }} }},
            grid: {{ color: CHART_GRID }},
            title: {{ display: true, text: '€/L per $10 Brent', color: '#e2e8f0', font: {{ size: 13, weight: '600' }} }},
          }},
        }},
      }},
      plugins: [{{
        id: 'sensLabels',
        afterDatasetsDraw(chart) {{
          const ctx2 = chart.ctx;
          chart.data.datasets.forEach((ds, i) => {{
            chart.getDatasetMeta(i).data.forEach((bar, j) => {{
              const val = ds.data[j];
              if (val == null || val === 0) return;
              ctx2.save();
              ctx2.fillStyle = '#e2e8f0';
              ctx2.font = 'bold 10px DM Sans, sans-serif';
              ctx2.textAlign = 'center';
              ctx2.textBaseline = 'bottom';
              ctx2.fillText('+' + (val/100).toFixed(2) + '€', bar.x, bar.y - 3);
              ctx2.restore();
            }});
          }});
        }}
      }}],
    }});
  }}

  makeBarChart('sensDieselChart', 'sens-diesel-slope', 'diesel');
  makeBarChart('sens95Chart',     'sens-95-slope',     'euro95');

  // Asymmetry tables (one per fuel, aligned under each chart)
  const sensD  = DATA.sensitivity.diesel.per_country;
  const sens95 = DATA.sensitivity.euro95.per_country;
  const fmtEuro = v => (v >= 0 ? '+' : '') + (v/100).toFixed(2) + '€';

  function buildAsymTable(wrapperId, sensData, upLabel, dnLabel) {{
    const wrap = $(wrapperId);
    const tbl = document.createElement('table');
    tbl.style.cssText = 'width:100%;border-collapse:collapse;font-size:12px;';
    tbl.innerHTML = `<thead><tr>
      <th style="text-align:left;padding:8px 16px;color:#94a3b8;font-weight:600;border-bottom:1px solid #1e293b;">Country</th>
      <th style="text-align:right;padding:8px 12px;color:#94a3b8;font-weight:600;border-bottom:1px solid #1e293b;">${{upLabel}}</th>
      <th style="text-align:right;padding:8px 12px;color:#94a3b8;font-weight:600;border-bottom:1px solid #1e293b;">${{dnLabel}}</th>
      <th style="text-align:right;padding:8px 12px;color:#94a3b8;font-weight:600;border-bottom:1px solid #1e293b;">Gap</th>
    </tr></thead><tbody></tbody>`;
    const tbody = tbl.querySelector('tbody');
    CTRS.filter(c => c !== 'EU').forEach(c => {{
      const up = sensData[c].up; const dn = sensData[c].down; const gap = up - dn;
      const tr = document.createElement('tr');
      tr.innerHTML = `
        <td style="padding:6px 16px;color:${{COLORS[c]}};font-weight:700;">${{c}}</td>
        <td class="mono" style="text-align:right;padding:6px 12px;color:#e2e8f0;">${{fmtEuro(up)}}</td>
        <td class="mono" style="text-align:right;padding:6px 12px;color:#e2e8f0;">${{fmtEuro(dn)}}</td>
        <td class="mono" style="text-align:right;padding:6px 12px;color:#e2e8f0;font-weight:700;">${{fmtEuro(gap)}}</td>`;
      tbody.appendChild(tr);
    }});
    wrap.appendChild(tbl);
  }}
  buildAsymTable('sens-diesel-table', sensD,  'Brent ▲', 'Brent ▼');
  buildAsymTable('sens-95-table',     sens95, 'Brent ▲', 'Brent ▼');

  // ---- Window × Lag research grid (data pre-computed in Python) ----------
  function buildSensResearch() {{
    const WINS = [4, 26];
    const LAGS = [0, 1, 2];

    function fillResearchTable(wrapperId, fuelKey) {{
      const research = DATA.sensitivity[fuelKey].research;
      const wrap = $(wrapperId);
      const tbl  = document.createElement('table');
      tbl.style.cssText = 'width:100%;border-collapse:collapse;font-size:12px;';
      tbl.innerHTML = `<thead><tr>
        <th style="padding:6px 12px;color:#94a3b8;font-weight:600;border-bottom:1px solid #1e293b;text-align:center;">Window</th>
        <th style="padding:6px 12px;color:#94a3b8;font-weight:600;border-bottom:1px solid #1e293b;text-align:center;">Lag</th>
        <th style="padding:6px 12px;color:#f59e0b;font-weight:600;border-bottom:1px solid #1e293b;text-align:right;">Brent ▲</th>
        <th style="padding:6px 12px;color:#64748b;font-weight:600;border-bottom:1px solid #1e293b;text-align:right;">Brent ▼</th>
        <th style="padding:6px 12px;color:#94a3b8;font-weight:600;border-bottom:1px solid #1e293b;text-align:right;">Gap</th>
      </tr></thead><tbody></tbody>`;
      const tbody = tbl.querySelector('tbody');
      WINS.forEach((w, wi) => {{
        if (wi > 0) {{
          const sep = document.createElement('tr');
          sep.innerHTML = '<td colspan="5" style="padding:0;border-top:2px solid #334155;"></td>';
          tbody.appendChild(sep);
        }}
        LAGS.forEach(l => {{
          const s         = research[`${{w}}_${{l}}`];
          const gap       = s.up - s.down;
          const isCurrent = (w === 4 && l === 1);
          const tr = document.createElement('tr');
          if (isCurrent) tr.style.cssText = 'background:rgba(245,158,11,0.12);';
          tr.innerHTML = `
            <td style="text-align:center;padding:5px 12px;color:#e2e8f0;font-weight:${{isCurrent?700:400}};">
              ${{w}}W${{isCurrent?' ★':''}}
            </td>
            <td style="text-align:center;padding:5px 12px;color:#94a3b8;">lag ${{l}}W</td>
            <td class="mono" style="text-align:right;padding:5px 12px;color:#f59e0b;">${{fmtEuro(s.up)}}</td>
            <td class="mono" style="text-align:right;padding:5px 12px;color:#64748b;">${{fmtEuro(s.down)}}</td>
            <td class="mono" style="text-align:right;padding:5px 12px;color:#e2e8f0;font-weight:700;">${{fmtEuro(gap)}}</td>`;
          tbody.appendChild(tr);
        }});
      }});
      wrap.appendChild(tbl);
    }}

    fillResearchTable('sens-research-diesel', 'diesel');
    fillResearchTable('sens-research-95',     'euro95');
  }}
  buildSensResearch();
}}
</script>
{analytics_html}
</body>
</html>"""


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
WEBPAGE_NAME  = "oil_dashboard"
GHPAGES_REPO  = Path(__file__).resolve().parent.parent / "sebast759.github.io"
GHPAGES_REMOTE = "https://github.com/sebast759/sebast759.github.io.git"
DEFAULT_OUT   = str(Path(__file__).resolve().parent / "site" / "index.html")


def push_to_github_pages(html: str, token: str):
    """Push HTML directly to GitHub Pages via the GitHub Contents API.

    No local git repo required — works on Heroku's ephemeral filesystem.

    Env vars (all optional except GITHUB_TOKEN):
      GITHUB_TOKEN     — personal access token with repo write scope
      GITHUB_REPO      — owner/repo  (default: sebast759/sebast759.github.io)
      GITHUB_FILE_PATH — path in repo (default: oil_dashboard/index.html)
    """
    import urllib.request
    import urllib.error

    repo      = os.environ.get("GITHUB_REPO",      "sebast759/sebast759.github.io")
    file_path = os.environ.get("GITHUB_FILE_PATH", "oil_dashboard/index.html")
    api_url   = f"https://api.github.com/repos/{repo}/contents/{file_path}"
    headers   = {
        "Authorization": f"token {token}",
        "Accept":        "application/vnd.github.v3+json",
        "Content-Type":  "application/json",
        "User-Agent":    "oil-dashboard-scheduler",
    }

    # Fetch current file SHA — required by the API for updates (not for first push)
    sha = None
    try:
        with urllib.request.urlopen(
            urllib.request.Request(api_url, headers=headers)
        ) as resp:
            sha = json.loads(resp.read())["sha"]
    except urllib.error.HTTPError as e:
        if e.code != 404:
            raise

    payload = {
        "message": f"chore: update oil dashboard ({datetime.now().strftime('%Y-%m-%d %H:%M')})",
        "content": base64.b64encode(html.encode()).decode(),
    }
    if sha:
        payload["sha"] = sha

    req = urllib.request.Request(
        api_url,
        data=json.dumps(payload).encode(),
        headers=headers,
        method="PUT",
    )
    with urllib.request.urlopen(req):
        pass

    owner = repo.split("/")[0]
    print(f"  Live at https://{owner}.github.io/{file_path.split('/')[0]}/")


def git_push(ghpages_repo: Path, webpage_name: str, source_file: Path):
    """Publish through a clean temporary checkout, preserving local Pages edits."""
    import subprocess
    print("\nPushing to GitHub Pages ...")
    generated_html = source_file.read_bytes()

    with tempfile.TemporaryDirectory(prefix="oil-dashboard-pages-") as tmp:
        clean_repo = Path(tmp) / "website"
        subprocess.run(
            ["git", "clone", "--quiet", "--single-branch",
             GHPAGES_REMOTE, str(clean_repo)],
            check=True,
        )

        target_file = clean_repo / webpage_name / "index.html"
        target_file.parent.mkdir(parents=True, exist_ok=True)
        target_file.write_bytes(generated_html)

        subprocess.run(["git", "config", "user.name", "Dashboard Bot"],
                       cwd=clean_repo, check=True)
        subprocess.run(["git", "config", "user.email", "actions@github.com"],
                       cwd=clean_repo, check=True)
        subprocess.run(["git", "add", f"{webpage_name}/index.html"],
                       cwd=clean_repo, check=True)
        result = subprocess.run(["git", "diff", "--cached", "--quiet"],
                                cwd=clean_repo)
        if result.returncode == 0:
            print("  Dashboard is already up to date.")
            return

        subprocess.run(
            ["git", "commit", "-m", f"Update {webpage_name} dashboard"],
            cwd=clean_repo, check=True,
        )
        subprocess.run(["git", "push"], cwd=clean_repo, check=True)

    print(f"  Live at https://sebast759.github.io/{webpage_name}/")


def main():
    parser = argparse.ArgumentParser(
        description="Generate EU Oil Bulletin Dashboard",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python generate_oil_dashboard.py                  # refresh data and create ./site/index.html
  python generate_oil_dashboard.py --download       # force re-download from EC website
  python generate_oil_dashboard.py my_file.xlsx     # use a specific local file
  python generate_oil_dashboard.py --local          # skip all network calls, use cache
  python generate_oil_dashboard.py --push           # generate and publish to GitHub Pages

Output (default):
  """ + DEFAULT_OUT
    )
    parser.add_argument("input", nargs="?", default=DEFAULT_XLSX,
                        help="Path to xlsx (optional — auto-downloads if omitted)")
    parser.add_argument("--output", "-o", default=DEFAULT_OUT,
                        help=f"Output HTML file (default: {DEFAULT_OUT})")
    parser.add_argument("--download", "-d", action="store_true",
                        help="Force re-download the latest file from EC website")
    parser.add_argument("--cache-dir", default=".",
                        help="Directory for cached xlsx (default: current folder)")
    parser.add_argument("--push", action="store_true",
                        help="Publish after generating (default: local file only)")
    parser.add_argument("--no-push", action="store_true",
                        help=argparse.SUPPRESS)
    parser.add_argument("--local", action="store_true",
                        help="Skip all network calls; use cached xlsx and brent_cache.csv")
    args = parser.parse_args()
    if args.push and args.no_push:
        parser.error("--push and --no-push cannot be used together")
    if args.push and args.local:
        parser.error("--push cannot be combined with --local")

    xlsx_path = resolve_xlsx(args.input, args.download, Path(args.cache_dir), local=args.local)

    data = extract_data(xlsx_path, local=args.local)
    html = build_html(data)

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(html, encoding="utf-8")
    emit_site_support_files(out_path.parent)
    print(f"\n  Dashboard saved to: {out_path.resolve()}")

    if args.push:
        github_token = os.environ.get("PAGES_TOKEN")
        if github_token:
            print("\nPushing to GitHub Pages via API ...")
            push_to_github_pages(html, github_token)
        elif GHPAGES_REPO.exists():
            git_push(GHPAGES_REPO, WEBPAGE_NAME, out_path)
        else:
            parser.error(
                f"cannot push: Pages repo not found at {GHPAGES_REPO} "
                "and PAGES_TOKEN is not set"
            )


if __name__ == "__main__":
    main()
